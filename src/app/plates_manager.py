"""
Interface graphique de gestion des plaques CORELAC - VERSION OPTIMISEE

- Cache SQLite pour chargement rapide
- Synchronisation automatique Excel <-> SQLite
"""

import os
import re
import sqlite3
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import shutil

# ============ CONFIGURATION ============
PLATES_FOLDER = r"C:\IE\Etudes\ET_Corélac\CORELAC_300_plaques_24_femelles_groupées\plaques_modifiees"
MODIFIED_FOLDER = os.path.join(os.path.dirname(PLATES_FOLDER), "plaques_modifiees")
os.makedirs(MODIFIED_FOLDER, exist_ok=True)

# Base de données cache
DB_PATH = os.path.join(MODIFIED_FOLDER, "corelac_cache.db")

# Dates de fertilisation
DATE_BOURGET = "19/12/2025"
DATE_LEMAN = "22/12/2025"

# Grille
ROWS = ['A', 'B', 'C', 'D']
COLUMNS = [1, 2, 3, 4, 5, 6]

# Couleurs
COLOR_ALIVE = '#00FF00'
COLOR_DEAD = '#FF0000'
COLOR_DEAD_EYED = '#FF6B35'
COLOR_DEAD_LARVAE = '#8B0000'
COLOR_RUNAWAY = '#9932CC'
COLOR_HOVER = '#FFFF99'
COLOR_HATCHED = '#CCFF00'
COLOR_HATCHED_TODAY = "#05F06F"
COLOR_HATCHED_14DAYS = '#A9A9A9'
COLOR_OUT_OF_STUDIES = '#505050'  # Gris foncé pour "Out of Studies"

# Pattern pour détecter les croisements L_M x L_F avec mâle 11-15
MALE_11_15_PATTERN = re.compile(r'L_M\s*(1[1-5])\s*x\s*L_F', re.IGNORECASE)

# Auges selon température
TANKS_5C = ['A1', 'A2']
TANKS_9C = ['A3', 'A4']
AUTO_TANK_LOGIC = True


# ============ CACHE SQLITE ============
class PlateCache:
    """Gestionnaire de cache SQLite pour les plaques"""

    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self.conn = None
        self._init_db()

    def _init_db(self):
        """Initialise la base de données"""
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row

        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS plates (
                plate_num INTEGER PRIMARY KEY,
                female_type TEXT,
                fert_date TEXT,
                last_sync TEXT
            );

            CREATE TABLE IF NOT EXISTS cells (
                plate_num INTEGER,
                position TEXT,
                cross_info TEXT,
                status TEXT,
                death_date TEXT,
                death_type TEXT,
                hatching_date TEXT,
                eyespot_date TEXT,
                photo_dates TEXT,
                last_photo_date TEXT,
                PRIMARY KEY (plate_num, position)
            );

            CREATE INDEX IF NOT EXISTS idx_cells_plate ON cells(plate_num);
            CREATE INDEX IF NOT EXISTS idx_cells_hatching ON cells(hatching_date);

            CREATE TABLE IF NOT EXISTS ignored_photo_conflicts (
                plate_num INTEGER,
                position TEXT,
                PRIMARY KEY (plate_num, position)
            );
        """)
        self.conn.commit()

    def is_plate_cached(self, plate_num):
        """Vérifie si une plaque est en cache"""
        cursor = self.conn.execute(
            "SELECT 1 FROM plates WHERE plate_num = ?", (plate_num,)
        )
        return cursor.fetchone() is not None

    def get_plate_data(self, plate_num):
        """Récupère les données d'une plaque depuis le cache"""
        # Info plaque
        cursor = self.conn.execute(
            "SELECT * FROM plates WHERE plate_num = ?", (plate_num,)
        )
        plate_row = cursor.fetchone()
        if not plate_row:
            return None

        # Cellules
        cursor = self.conn.execute(
            "SELECT * FROM cells WHERE plate_num = ? ORDER BY position",
            (plate_num,)
        )
        cells = {}
        for row in cursor.fetchall():
            photo_dates = row['photo_dates'].split('|') if row['photo_dates'] else []
            cells[row['position']] = {
                'cross': row['cross_info'] or '',
                'alive': row['status'] != 'Dead',
                'death_date': row['death_date'] or '',
                'death_type': row['death_type'] or 'Dead',
                'hatching_date': row['hatching_date'] or '',
                'eyespot_date': row['eyespot_date'] or '',
                'photo_dates': photo_dates,
                'last_photo_date': row['last_photo_date'] or ''
            }

        return {
            'female_type': plate_row['female_type'],
            'fert_date': plate_row['fert_date'],
            'cells': cells
        }

    def save_plate_to_cache(self, plate_num, female_type, fert_date, cells_status):
        """Sauvegarde une plaque dans le cache"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Upsert plaque
        self.conn.execute("""
            INSERT OR REPLACE INTO plates (plate_num, female_type, fert_date, last_sync)
            VALUES (?, ?, ?, ?)
        """, (plate_num, female_type, fert_date, now))

        # Upsert cellules
        for pos, data in cells_status.items():
            photo_dates_str = '|'.join(data.get('photo_dates', []))
            status = 'Alive' if data.get('alive', True) else 'Dead'

            self.conn.execute("""
                INSERT OR REPLACE INTO cells
                (plate_num, position, cross_info, status, death_date, death_type,
                 hatching_date, eyespot_date, photo_dates, last_photo_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                plate_num, pos, data.get('cross', ''), status,
                data.get('death_date', ''), data.get('death_type', 'Dead'),
                data.get('hatching_date', ''), data.get('eyespot_date', ''),
                photo_dates_str, data.get('last_photo_date', '')
            ))

        self.conn.commit()

    def update_cell(self, plate_num, position, **kwargs):
        """Met à jour une cellule spécifique"""
        # Construire la requête UPDATE dynamiquement
        updates = []
        values = []

        field_mapping = {
            'cross': 'cross_info',
            'alive': 'status',
            'death_date': 'death_date',
            'death_type': 'death_type',
            'hatching_date': 'hatching_date',
            'eyespot_date': 'eyespot_date',
            'photo_dates': 'photo_dates',
            'last_photo_date': 'last_photo_date'
        }

        for key, value in kwargs.items():
            if key in field_mapping:
                db_field = field_mapping[key]
                if key == 'alive':
                    value = 'Alive' if value else 'Dead'
                elif key == 'photo_dates':
                    value = '|'.join(value) if isinstance(value, list) else value
                updates.append(f"{db_field} = ?")
                values.append(value)

        if updates:
            values.extend([plate_num, position])
            self.conn.execute(f"""
                UPDATE cells SET {', '.join(updates)}
                WHERE plate_num = ? AND position = ?
            """, values)
            self.conn.commit()

    def get_ignored_conflicts(self, plate_num):
        """Récupère les positions des conflits ignorés pour une plaque"""
        cursor = self.conn.execute("""
            SELECT position FROM ignored_photo_conflicts
            WHERE plate_num = ?
        """, (plate_num,))
        return {row['position'] for row in cursor.fetchall()}

    def add_ignored_conflict(self, plate_num, position):
        """Ajoute un conflit ignoré"""
        self.conn.execute("""
            INSERT OR IGNORE INTO ignored_photo_conflicts (plate_num, position)
            VALUES (?, ?)
        """, (plate_num, position))
        self.conn.commit()

    def remove_ignored_conflict(self, plate_num, position):
        """Retire un conflit ignoré"""
        self.conn.execute("""
            DELETE FROM ignored_photo_conflicts
            WHERE plate_num = ? AND position = ?
        """, (plate_num, position))
        self.conn.commit()

    def get_plates_needing_photos(self):
        """Retourne les plaques avec des alevins à photographier"""
        today = datetime.now()
        three_days_ago = today.strftime("%d/%m/%Y")

        cursor = self.conn.execute("""
            SELECT DISTINCT plate_num FROM cells
            WHERE status != 'Dead'
            AND hatching_date IS NOT NULL
            AND hatching_date != ''
            ORDER BY plate_num
        """)

        plates_with_photos = []
        for row in cursor.fetchall():
            plate_num = row['plate_num']
            # Vérifier si au moins une cellule a besoin de photo
            cells_cursor = self.conn.execute("""
                SELECT position, hatching_date, last_photo_date FROM cells
                WHERE plate_num = ? AND status != 'Dead' AND hatching_date IS NOT NULL
            """, (plate_num,))

            for cell in cells_cursor.fetchall():
                if self._needs_photo_check(cell['hatching_date'], cell['last_photo_date'], today):
                    plates_with_photos.append(plate_num)
                    break

        return plates_with_photos

    def _needs_photo_check(self, hatching_str, last_photo_str, today):
        """Vérifie si une cellule a besoin de photo"""
        if not hatching_str:
            return False

        try:
            if not last_photo_str:
                hatching_date = datetime.strptime(hatching_str, "%d/%m/%Y")
                return (today - hatching_date).days >= 3
            else:
                last_photo_date = datetime.strptime(last_photo_str, "%d/%m/%Y")
                return (today - last_photo_date).days >= 3
        except:
            return False

    def close(self):
        """Ferme la connexion"""
        if self.conn:
            self.conn.close()


# ============ SYNCHRONISATION EXCEL ============
class ExcelSyncer:
    """Synchronise le cache avec les fichiers Excel"""

    @staticmethod
    def sync_plate_to_cache(plate_num, cache: PlateCache, progress_callback=None):
        """Charge une plaque Excel et la met en cache"""
        plate_name = f"Plaque_{plate_num:03d}.xlsx"

        # Chercher le fichier
        plate_path = os.path.join(MODIFIED_FOLDER, plate_name)
        if not os.path.exists(plate_path):
            plate_path = os.path.join(PLATES_FOLDER, plate_name)

        if not os.path.exists(plate_path):
            return False

        try:
            # Charger en mode lecture seule pour plus de rapidité
            wb = openpyxl.load_workbook(plate_path, read_only=True, data_only=True)

            ws_disp = wb["Disposition"]
            ws_suivi = wb["Suivi"]

            # Détecter type de femelle
            first_cross = None
            for row_idx in range(len(ROWS)):
                for col_idx in range(len(COLUMNS)):
                    cell_val = ws_disp.cell(row_idx + 2, col_idx + 2).value
                    if cell_val and 'x' in str(cell_val):
                        first_cross = str(cell_val)
                        break
                if first_cross:
                    break

            if first_cross and 'B_F' in first_cross:
                fert_date = DATE_BOURGET
                female_type = "Bourget"
            elif first_cross and 'L_F' in first_cross:
                fert_date = DATE_LEMAN
                female_type = "Léman"
            else:
                fert_date = "Inconnue"
                female_type = "?"

            # Charger les cellules
            cells_status = {}

            for row_idx, row in enumerate(ROWS):
                for col_idx, col in enumerate(COLUMNS):
                    pos = f"{row}{col}"

                    cross = ws_disp.cell(row_idx + 2, col_idx + 2).value
                    cross = str(cross).strip() if cross else ""

                    suivi_row = row_idx * len(COLUMNS) + col_idx + 2
                    status = ws_suivi.cell(suivi_row, 9).value
                    death_date = ws_suivi.cell(suivi_row, 10).value
                    death_type = ws_suivi.cell(suivi_row, 11).value
                    hatching_date = ws_suivi.cell(suivi_row, 8).value
                    eyespot_date = ws_suivi.cell(suivi_row, 7).value

                    # Charger les dates de photos
                    photo_dates = []
                    for photo_col in range(13, 33):
                        photo_date = ws_suivi.cell(suivi_row, photo_col).value
                        if photo_date and str(photo_date).strip():
                            photo_dates.append(str(photo_date).strip())
                        else:
                            break

                    # Calculer dernière date photo
                    last_photo_date_str = ""
                    if photo_dates:
                        valid_dates = []
                        for date_str in photo_dates:
                            try:
                                date_obj = datetime.strptime(date_str, "%d/%m/%Y")
                                valid_dates.append((date_obj, date_str))
                            except:
                                try:
                                    date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                                    valid_dates.append((date_obj, date_obj.strftime("%d/%m/%Y")))
                                except:
                                    pass

                        if valid_dates:
                            valid_dates.sort(key=lambda x: x[0], reverse=True)
                            last_photo_date_str = valid_dates[0][1]

                    cells_status[pos] = {
                        'cross': cross,
                        'alive': status != "Dead",
                        'death_date': str(death_date).strip() if death_date else "",
                        'death_type': str(death_type).strip() if death_type else "Dead",
                        'hatching_date': str(hatching_date).strip() if hatching_date else "",
                        'eyespot_date': str(eyespot_date).strip() if eyespot_date else "",
                        'photo_dates': photo_dates,
                        'last_photo_date': last_photo_date_str
                    }

            wb.close()

            # Sauvegarder dans le cache
            cache.save_plate_to_cache(plate_num, female_type, fert_date, cells_status)

            if progress_callback:
                progress_callback(plate_num)

            return True

        except Exception as e:
            print(f"Erreur sync plaque {plate_num}: {e}")
            return False

    @staticmethod
    def sync_all_plates(cache: PlateCache, progress_callback=None, complete_callback=None):
        """Synchronise toutes les plaques en arrière-plan"""
        def worker():
            synced = 0
            for plate_num in range(1, 301):
                if ExcelSyncer.sync_plate_to_cache(plate_num, cache, progress_callback):
                    synced += 1

            if complete_callback:
                complete_callback(synced)

        thread = threading.Thread(target=worker, daemon=True)
        thread.start()
        return thread


# ============ COMPLETE CELL EDITOR ============
class CompleteCellEditor:
    """Complete cell data editor - Modal dialog"""

    def __init__(self, parent, plate_manager, initial_pos="A1"):
        self.parent = parent
        self.pm = plate_manager  # Référence au PlateManagerFast
        self.current_pos = initial_pos
        self.original_data = {}  # Pour la fonction Reset

        # Créer le dialogue
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Complete Cell Data Editor - {initial_pos}")
        self.dialog.geometry("550x700")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Centrer
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")

        # Variables
        self.position_var = tk.StringVar(value=initial_pos)
        self.status_var = tk.StringVar(value="Alive")
        self.death_date_var = tk.StringVar()
        self.death_type_var = tk.StringVar(value="Dead")
        self.eyespot_date_var = tk.StringVar()
        self.hatching_date_var = tk.StringVar()

        # Construire l'interface
        self._setup_ui()

        # Charger les données initiales
        self._load_position_data(initial_pos)

        # Raccourcis clavier
        self.dialog.bind('<Escape>', lambda e: self.dialog.destroy())
        self.dialog.bind('<Return>', lambda e: self._validate_and_save())

    def _setup_ui(self):
        """Construit l'interface de l'éditeur"""
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ===== POSITION SELECTOR =====
        pos_frame = ttk.LabelFrame(main_frame, text="Cell Position", padding="10")
        pos_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(pos_frame, text="Position:").pack(side=tk.LEFT, padx=5)

        # Dropdown avec toutes les positions
        positions = [f"{row}{col}" for row in ROWS for col in COLUMNS]
        pos_combo = ttk.Combobox(pos_frame, textvariable=self.position_var,
                                  values=positions, width=8, state='readonly')
        pos_combo.pack(side=tk.LEFT, padx=5)
        pos_combo.bind('<<ComboboxSelected>>', lambda e: self._on_position_changed())

        # Boutons de navigation
        ttk.Button(pos_frame, text="◀", width=3,
                   command=self._navigate_previous).pack(side=tk.LEFT, padx=2)
        ttk.Button(pos_frame, text="▶", width=3,
                   command=self._navigate_next).pack(side=tk.LEFT, padx=2)

        # Label du croisement
        self.cross_label = tk.Label(pos_frame, text="", font=("Arial", 9), fg="blue")
        self.cross_label.pack(side=tk.LEFT, padx=10)

        # ===== STATUS =====
        status_frame = ttk.LabelFrame(main_frame, text="Status", padding="10")
        status_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Radiobutton(status_frame, text="Alive", variable=self.status_var,
                        value="Alive", command=self._on_status_changed).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(status_frame, text="Dead", variable=self.status_var,
                        value="Dead", command=self._on_status_changed).pack(side=tk.LEFT, padx=5)

        # ===== DEATH DATA =====
        death_frame = ttk.LabelFrame(main_frame, text="Death Data", padding="10")
        death_frame.pack(fill=tk.X, pady=(0, 10))

        # Death Date
        date_frame = ttk.Frame(death_frame)
        date_frame.pack(fill=tk.X, pady=5)
        ttk.Label(date_frame, text="Death Date (DD/MM/YYYY):").pack(side=tk.LEFT, padx=5)
        self.death_date_entry = ttk.Entry(date_frame, textvariable=self.death_date_var, width=15)
        self.death_date_entry.pack(side=tk.LEFT, padx=5)

        # Death Type
        type_frame = ttk.Frame(death_frame)
        type_frame.pack(fill=tk.X, pady=5)
        ttk.Label(type_frame, text="Death Type:").pack(side=tk.LEFT, padx=5)
        death_types = ["Dead", "Dead eyed", "Dead larvae", "Runaway", "Out of Studies"]
        self.death_type_combo = ttk.Combobox(type_frame, textvariable=self.death_type_var,
                                              values=death_types, width=20, state='readonly')
        self.death_type_combo.pack(side=tk.LEFT, padx=5)

        # ===== DEVELOPMENT DATES =====
        dev_frame = ttk.LabelFrame(main_frame, text="Development Dates", padding="10")
        dev_frame.pack(fill=tk.X, pady=(0, 10))

        # Eyespot Date
        eye_frame = ttk.Frame(dev_frame)
        eye_frame.pack(fill=tk.X, pady=5)
        ttk.Label(eye_frame, text="Eyespot Date:").pack(side=tk.LEFT, padx=5)
        self.eyespot_entry = ttk.Entry(eye_frame, textvariable=self.eyespot_date_var, width=15)
        self.eyespot_entry.pack(side=tk.LEFT, padx=5)

        # Hatching Date
        hatch_frame = ttk.Frame(dev_frame)
        hatch_frame.pack(fill=tk.X, pady=5)
        ttk.Label(hatch_frame, text="Hatching Date:").pack(side=tk.LEFT, padx=5)
        self.hatching_entry = ttk.Entry(hatch_frame, textvariable=self.hatching_date_var, width=15)
        self.hatching_entry.pack(side=tk.LEFT, padx=5)

        # ===== PHOTO DATES =====
        photo_frame = ttk.LabelFrame(main_frame, text="Photo Dates (Max 20)", padding="10")
        photo_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Liste des photos avec scrollbar
        list_frame = ttk.Frame(photo_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.photo_listbox = tk.Listbox(list_frame, height=8, yscrollcommand=scrollbar.set)
        self.photo_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.photo_listbox.yview)

        # Boutons pour gérer les photos
        photo_btn_frame = ttk.Frame(photo_frame)
        photo_btn_frame.pack(fill=tk.X, pady=(5, 0))

        self.photo_date_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        ttk.Entry(photo_btn_frame, textvariable=self.photo_date_var, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(photo_btn_frame, text="+ Add", command=self._add_photo_date).pack(side=tk.LEFT, padx=2)
        ttk.Button(photo_btn_frame, text="- Remove", command=self._remove_photo_date).pack(side=tk.LEFT, padx=2)

        # ===== BUTTONS =====
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(btn_frame, text="Save", command=self._validate_and_save).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Reset", command=self._reset_form).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.LEFT, padx=5)

    def _load_position_data(self, pos):
        """Charge les données pour une position donnée"""
        if pos not in self.pm.cells_status:
            return

        data = self.pm.cells_status[pos]
        self.original_data = dict(data)  # Copie pour reset

        # Mettre à jour le titre
        self.dialog.title(f"Complete Cell Data Editor - {pos}")

        # Afficher le croisement
        cross = data.get('cross', '')
        self.cross_label.config(text=f"Cross: {cross}")

        # Status
        if data.get('alive', True):
            self.status_var.set("Alive")
        else:
            self.status_var.set("Dead")

        # Death data
        self.death_date_var.set(data.get('death_date', ''))
        self.death_type_var.set(data.get('death_type', 'Dead'))

        # Development dates
        self.eyespot_date_var.set(data.get('eyespot_date', ''))
        self.hatching_date_var.set(data.get('hatching_date', ''))

        # Photo dates
        self.photo_listbox.delete(0, tk.END)
        photo_dates = data.get('photo_dates', [])
        for i, pd in enumerate(photo_dates):
            self.photo_listbox.insert(tk.END, f"{i+1}. {pd}")

        # Activer/désactiver les champs selon le statut
        self._on_status_changed()

    def _on_position_changed(self):
        """Appelé quand l'utilisateur change de position"""
        new_pos = self.position_var.get()
        self._load_position_data(new_pos)
        self.current_pos = new_pos

    def _navigate_previous(self):
        """Navigue vers la cellule précédente"""
        positions = [f"{row}{col}" for row in ROWS for col in COLUMNS]
        current_idx = positions.index(self.current_pos)
        if current_idx > 0:
            new_pos = positions[current_idx - 1]
            self.position_var.set(new_pos)
            self._load_position_data(new_pos)
            self.current_pos = new_pos

    def _navigate_next(self):
        """Navigue vers la cellule suivante"""
        positions = [f"{row}{col}" for row in ROWS for col in COLUMNS]
        current_idx = positions.index(self.current_pos)
        if current_idx < len(positions) - 1:
            new_pos = positions[current_idx + 1]
            self.position_var.set(new_pos)
            self._load_position_data(new_pos)
            self.current_pos = new_pos

    def _on_status_changed(self):
        """Active/désactive les champs selon le statut"""
        status = self.status_var.get()

        if status == "Alive":
            # Désactiver les champs de mort
            self.death_date_entry.config(state=tk.DISABLED)
            self.death_type_combo.config(state=tk.DISABLED)
        else:
            # Activer les champs de mort
            self.death_date_entry.config(state=tk.NORMAL)
            self.death_type_combo.config(state='readonly')

    def _add_photo_date(self):
        """Ajoute une date de photo"""
        if self.photo_listbox.size() >= 20:
            messagebox.showwarning("Limite atteinte", "Maximum 20 dates de photos par cellule")
            return

        photo_date = self.photo_date_var.get().strip()
        if not photo_date:
            return

        # Valider le format
        try:
            datetime.strptime(photo_date, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erreur", "Format de date invalide.\nUtilisez DD/MM/YYYY (ex: 15/03/2026)")
            return

        # Vérifier si déjà présente
        current_dates = [self.photo_listbox.get(i).split(". ", 1)[1] for i in range(self.photo_listbox.size())]
        if photo_date in current_dates:
            messagebox.showinfo("Info", "Cette date est déjà dans la liste")
            return

        # Ajouter
        idx = self.photo_listbox.size() + 1
        self.photo_listbox.insert(tk.END, f"{idx}. {photo_date}")

    def _remove_photo_date(self):
        """Supprime la date de photo sélectionnée"""
        selection = self.photo_listbox.curselection()
        if not selection:
            messagebox.showinfo("Info", "Sélectionnez une date à supprimer")
            return

        self.photo_listbox.delete(selection[0])

        # Re-numéroter
        dates = [self.photo_listbox.get(i).split(". ", 1)[1] for i in range(self.photo_listbox.size())]
        self.photo_listbox.delete(0, tk.END)
        for i, date in enumerate(dates):
            self.photo_listbox.insert(tk.END, f"{i+1}. {date}")

    def _validate_and_save(self):
        """Valide et sauvegarde les modifications"""
        # Validation des dates
        dates_to_validate = [
            (self.death_date_var.get(), "Death Date"),
            (self.eyespot_date_var.get(), "Eyespot Date"),
            (self.hatching_date_var.get(), "Hatching Date")
        ]

        for date_str, field_name in dates_to_validate:
            if date_str.strip():
                try:
                    datetime.strptime(date_str, "%d/%m/%Y")
                except ValueError:
                    messagebox.showerror("Erreur",
                        f"Format de date invalide pour {field_name}.\nUtilisez DD/MM/YYYY (ex: 15/03/2026)")
                    return

        # Validation logique
        status = self.status_var.get()
        if status == "Dead":
            if not self.death_date_var.get().strip():
                messagebox.showerror("Erreur", "La date de mort est obligatoire pour un statut 'Dead'")
                return

        # Sauvegarder
        pos = self.current_pos
        data = self.pm.cells_status[pos]

        # Status
        if status == "Alive":
            data['alive'] = True
            data['death_date'] = ''
            data['death_type'] = ''
        else:  # Dead
            data['alive'] = False
            data['death_date'] = self.death_date_var.get().strip()
            data['death_type'] = self.death_type_var.get()

        # Development dates
        data['eyespot_date'] = self.eyespot_date_var.get().strip()
        data['hatching_date'] = self.hatching_date_var.get().strip()

        # Photo dates
        photo_dates = [self.photo_listbox.get(i).split(". ", 1)[1] for i in range(self.photo_listbox.size())]
        data['photo_dates'] = photo_dates
        data['last_photo_date'] = photo_dates[-1] if photo_dates else ''

        # Mettre à jour le cache
        self.pm.cache.update_cell(
            self.pm.current_plate_number, pos,
            alive=data['alive'],
            death_date=data['death_date'],
            death_type=data['death_type'],
            eyespot_date=data['eyespot_date'],
            hatching_date=data['hatching_date'],
            photo_dates=photo_dates,
            last_photo_date=data['last_photo_date']
        )

        # Rafraîchir l'UI
        self.pm._refresh_button(pos)
        self.pm.update_status()
        self.pm.check_photo_death_conflicts()
        self.pm.update_photo_tracking()
        self.pm.mark_unsaved()

        # Sauvegarder en arrière-plan
        self.pm._save_excel_background()

        messagebox.showinfo("Succès", f"Données sauvegardées pour {pos}")

        # Charger les nouvelles données pour continuer l'édition
        self._load_position_data(pos)

    def _reset_form(self):
        """Réinitialise le formulaire aux valeurs originales"""
        self._load_position_data(self.current_pos)


# ============ PLATE MANAGER OPTIMISÉ ============
class PlateManagerFast:
    def __init__(self, root):
        self.root = root
        self.root.title("🧬 Gestion Plaques CORELAC [FAST]")
        self.root.geometry("1000x850")

        # Définir l'icône de la fenêtre (pour la barre des tâches Windows)
        try:
            import sys
            if getattr(sys, 'frozen', False):
                # Si exécuté depuis un .exe compilé, utiliser l'icône de l'exécutable
                self.root.iconbitmap(default=sys.executable)
            else:
                # En développement, essayer de charger l'icône depuis le fichier
                icon_path = os.path.join(
                    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                    "assets",
                    "icons",
                    "corelac_bimp_hq.ico",
                )
                if os.path.exists(icon_path):
                    self.root.iconbitmap(default=icon_path)
        except Exception as e:
            # Si l'icône ne peut pas être chargée, continuer sans
            pass

        # Cache SQLite
        self.cache = PlateCache()

        # Variables d'état
        self.current_plate = None
        self.current_plate_number = None
        self.wb = None  # Workbook Excel (pour sauvegarde)
        self.cells_status = {}
        self.selected_dead = set()
        self.buttons = {}
        self.newly_marked_dead = set()
        self.hatching_mode = False
        self.photo_mode = False
        self.blink_cells = set()
        self.ignored_photo_conflicts = set()
        self.has_unsaved_changes = False  # Tracker les modifications non sauvegardées

        # Variables pour le type de femelle
        self.current_female_type = ""
        self.current_fert_date = ""

        # Construire l'interface
        self.setup_ui()

        # Synchroniser le cache en arrière-plan
        self.start_background_sync()

    def start_background_sync(self):
        """Lance la synchronisation du cache en arrière-plan"""
        def on_progress(plate_num):
            # Mise à jour discrète
            pass

        def on_complete(synced):
            self.root.after(0, lambda: self.label_cache_status.config(
                text=f"✅ Cache: {synced} plaques", fg="green"
            ))

        self.label_cache_status.config(text="🔄 Sync cache...", fg="orange")
        ExcelSyncer.sync_all_plates(self.cache, on_progress, on_complete)

    def setup_ui(self):
        """Construit l'interface"""
        # Setup menu bar
        self._setup_menu_bar()

        # Frame supérieure
        frame_top = ttk.Frame(self.root, padding="5")
        frame_top.pack(fill=tk.X)

        ttk.Label(frame_top, text="Plaque N°:").pack(side=tk.LEFT, padx=5)
        self.entry_plate = ttk.Entry(frame_top, width=10, font=("Arial", 12))
        self.entry_plate.pack(side=tk.LEFT, padx=5)
        self.entry_plate.bind('<Return>', lambda e: self.load_plate())

        ttk.Button(frame_top, text="Charger", command=self.load_plate).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_top, text="◀", command=self.load_previous_plate, width=3).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_top, text="▶", command=self.load_next_plate, width=3).pack(side=tk.LEFT, padx=2)

        # Indicateur modifications non sauvegardées
        self.label_unsaved = tk.Label(frame_top, text="", font=("Arial", 12, "bold"))
        self.label_unsaved.pack(side=tk.RIGHT, padx=5)

        # Status cache
        self.label_cache_status = tk.Label(frame_top, text="", font=("Arial", 9))
        self.label_cache_status.pack(side=tk.RIGHT, padx=10)

        # Bouton resync
        ttk.Button(frame_top, text="🔄 Resync", command=self.resync_current_plate).pack(side=tk.RIGHT, padx=5)

        # Frame dates
        frame_dates = ttk.Frame(self.root, padding="5")
        frame_dates.pack(fill=tk.X)

        ttk.Label(frame_dates, text="Fertilisation:").pack(side=tk.LEFT, padx=5)
        self.label_fert = tk.Label(frame_dates, text="---", font=("Arial", 10), fg="green")
        self.label_fert.pack(side=tk.LEFT, padx=5)

        ttk.Label(frame_dates, text="Eyespot:").pack(side=tk.LEFT, padx=20)
        self.entry_eyespot = ttk.Entry(frame_dates, width=12)
        self.entry_eyespot.pack(side=tk.LEFT, padx=5)
        self.entry_eyespot.insert(0, datetime.now().strftime("%d/%m/%Y"))

        ttk.Button(frame_dates, text="Appliquer Eyespot", command=self.apply_eyespot).pack(side=tk.LEFT, padx=5)

        # Frame éclosion
        frame_hatch = ttk.Frame(self.root, padding="5")
        frame_hatch.pack(fill=tk.X)

        ttk.Label(frame_hatch, text="Date éclosion:").pack(side=tk.LEFT, padx=5)
        self.entry_hatching = ttk.Entry(frame_hatch, width=12)
        self.entry_hatching.pack(side=tk.LEFT, padx=5)
        self.entry_hatching.insert(0, datetime.now().strftime("%d/%m/%Y"))

        self.btn_hatching = ttk.Button(frame_hatch, text="🐣 Mode éclosion (O)", command=self.toggle_hatching_mode)
        self.btn_hatching.pack(side=tk.LEFT, padx=5)

        # Frame photo
        frame_photo = ttk.Frame(self.root, padding="5")
        frame_photo.pack(fill=tk.X)

        self.btn_photo_mode = tk.Button(
            frame_photo, text="📷 Mode photo (P)",
            command=self.toggle_photo_mode, bg="#FF9800", fg="white"
        )
        self.btn_photo_mode.pack(side=tk.LEFT, padx=5)

        # Bouton appliquer photo à tous
        ttk.Button(frame_photo, text="📷 Photo → Tous vivants",
                   command=self.apply_photo_to_all_alive).pack(side=tk.LEFT, padx=5)

        self.label_photo_indicator = tk.Label(frame_photo, text="", font=("Arial", 10))
        self.label_photo_indicator.pack(side=tk.LEFT, padx=10)

        # Frame type de mort
        frame_death = ttk.Frame(self.root, padding="5")
        frame_death.pack(fill=tk.X)

        ttk.Label(frame_death, text="Type de mort:").pack(side=tk.LEFT, padx=5)
        self.death_type_var = tk.StringVar(value="Dead")

        for dtype, emoji in [("Dead", "☠️"), ("Dead eyed", "👁️"), ("Dead larvae", "🐟"), ("Runaway", "🏃"), ("Out of Studies", "📤")]:
            ttk.Radiobutton(
                frame_death, text=f"{emoji} {dtype}",
                variable=self.death_type_var, value=dtype
            ).pack(side=tk.LEFT, padx=5)

        # Bouton pour appliquer Out of Studies automatiquement
        ttk.Button(frame_death, text="📤 Auto Out of Studies (>14j)",
                   command=self.apply_out_of_studies).pack(side=tk.LEFT, padx=10)

        # Indicateurs
        frame_indicators = ttk.Frame(self.root, padding="5")
        frame_indicators.pack(fill=tk.X)

        self.label_eyespot_indicator = tk.Label(frame_indicators, text="", font=("Arial", 10))
        self.label_eyespot_indicator.pack(side=tk.LEFT, padx=10)

        self.label_next_photo = tk.Label(frame_indicators, text="", font=("Arial", 10, "bold"), fg="blue")
        self.label_next_photo.pack(side=tk.RIGHT, padx=10)

        # Grille
        frame_grid = ttk.Frame(self.root, padding="10")
        frame_grid.pack(fill=tk.BOTH, expand=True)

        # En-têtes colonnes
        for col_idx, col in enumerate(COLUMNS):
            lbl = ttk.Label(frame_grid, text=str(col), font=("Arial", 10, "bold"))
            lbl.grid(row=0, column=col_idx+1, padx=5)

        # En-têtes lignes + boutons
        for row_idx, row in enumerate(ROWS):
            lbl = ttk.Label(frame_grid, text=row, font=("Arial", 10, "bold"))
            lbl.grid(row=row_idx+1, column=0, padx=5)

            for col_idx, col in enumerate(COLUMNS):
                pos = f"{row}{col}"
                btn = tk.Button(
                    frame_grid, text="",
                    font=("Arial", 9), width=14, height=3,
                    relief=tk.RAISED, bd=3, bg=COLOR_ALIVE,
                    command=lambda p=pos: self.toggle_cell(p)
                )
                btn.grid(row=row_idx+1, column=col_idx+1, padx=2, pady=2, sticky="nsew")
                btn.bind('<Button-3>', lambda e, p=pos: self.show_context_menu(e, p))
                self.buttons[pos] = btn
                self.create_tooltip(btn, pos)

        # Configuration redimensionnement
        for i in range(len(COLUMNS) + 1):
            frame_grid.columnconfigure(i, weight=1)
        for i in range(len(ROWS) + 1):
            frame_grid.rowconfigure(i, weight=1)

        # Frame statut
        frame_status = ttk.Frame(self.root, padding="5")
        frame_status.pack(fill=tk.X)

        self.label_status = tk.Label(frame_status, text="Aucune plaque chargée", font=("Arial", 11))
        self.label_status.pack(side=tk.LEFT, padx=10)

        self.label_photo_death_warning = tk.Label(frame_status, text="", font=("Arial", 10, "bold"), fg="red")
        self.label_photo_death_warning.pack(side=tk.LEFT, padx=10)

        ttk.Button(frame_status, text="💾 Sauvegarder", command=self.save_plate).pack(side=tk.RIGHT, padx=10)

        # Bindings clavier
        self.root.bind('<Left>', lambda e: self._handle_arrow('left'))
        self.root.bind('<Right>', lambda e: self._handle_arrow('right'))
        self.root.bind('p', lambda e: self.toggle_photo_mode())
        self.root.bind('P', lambda e: self.toggle_photo_mode())
        self.root.bind('o', lambda e: self.toggle_hatching_mode())
        self.root.bind('O', lambda e: self.toggle_hatching_mode())

        # Keyboard shortcuts
        self.root.bind('<Control-e>', lambda e: self.open_complete_editor())
        self.root.bind('<Control-s>', lambda e: self.save_plate())
        self.root.bind('<Control-r>', lambda e: self.resync_current_plate())

    def _handle_arrow(self, direction):
        """Gère les touches flèches - navigation silencieuse"""
        focused = self.root.focus_get()
        if isinstance(focused, tk.Entry):
            return

        if direction == 'left':
            self.load_previous_plate_silent()
        else:
            self.load_next_plate_silent()

    def load_plate(self):
        """Charge une plaque - depuis le cache si disponible"""
        # Vérification sauvegarde plaque précédente
        if self.current_plate_number and self.cells_status:
            user_choice = self._show_save_dialog()

            if user_choice == 'cancel':
                return

            if user_choice == 'save':
                self._save_plate_silent()

        plate_str = self.entry_plate.get().strip()
        if not plate_str:
            return

        try:
            plate_num = int(plate_str)
        except ValueError:
            messagebox.showerror("Erreur", "Numéro de plaque invalide")
            return

        # Essayer le cache d'abord
        cached_data = self.cache.get_plate_data(plate_num)

        if cached_data:
            self._load_from_cache(plate_num, cached_data)
        else:
            # Synchroniser puis charger
            if ExcelSyncer.sync_plate_to_cache(plate_num, self.cache):
                cached_data = self.cache.get_plate_data(plate_num)
                if cached_data:
                    self._load_from_cache(plate_num, cached_data)
                    return

            messagebox.showerror("Erreur", f"Plaque {plate_num} introuvable")

    def _load_from_cache(self, plate_num, cached_data):
        """Charge les données depuis le cache"""
        # Charger les conflits ignorés depuis la DB (persistés)
        self.ignored_photo_conflicts = self.cache.get_ignored_conflicts(plate_num)

        self.current_plate_number = plate_num
        self.current_plate = f"Plaque_{plate_num:03d}.xlsx"
        self.current_female_type = cached_data['female_type']
        self.current_fert_date = cached_data['fert_date']

        self.cells_status = cached_data['cells']
        self.selected_dead = set()
        self.newly_marked_dead = set()

        # Mettre à jour l'affichage
        self.label_fert.config(text=f"{self.current_fert_date} ({self.current_female_type}) ✅")

        # Compter eyespot
        eyespot_count = 0
        eyespot_dates = {}

        for pos, data in self.cells_status.items():
            btn = self.buttons[pos]

            # Texte du bouton
            cross = data.get('cross', '')
            if MALE_11_15_PATTERN.search(cross):
                btn_text = f"⭕ {cross}" if len(cross) <= 12 else f"⭕ {cross[:10]}..."
            else:
                btn_text = cross if len(cross) <= 15 else cross[:12] + "..."

            # Indicateurs
            indicators = []
            if not data['alive'] and data['death_date']:
                emoji, _ = self.get_death_emoji_and_color(data['death_type'])
                indicators.append(f"{emoji} {data['death_date']}")
                self.selected_dead.add(pos)

            if data['hatching_date']:
                indicators.append(f"🐣 {data['hatching_date']}")

            if indicators:
                btn_text = f"{btn_text}\n" + "\n".join(indicators)

            btn.config(text=btn_text)

            # Couleur
            self._update_button_color(pos, data)

            # Eyespot
            if data.get('eyespot_date'):
                eyespot_count += 1
                date_str = data['eyespot_date']
                eyespot_dates[date_str] = eyespot_dates.get(date_str, 0) + 1

        # Indicateur eyespot
        if eyespot_count > 0:
            if len(eyespot_dates) == 1:
                self.label_eyespot_indicator.config(
                    text=f"✅ {eyespot_count}/24 œufs - Date: {list(eyespot_dates.keys())[0]}",
                    fg="green"
                )
            else:
                dates_summary = ", ".join([f"{d} ({c})" for d, c in eyespot_dates.items()])
                self.label_eyespot_indicator.config(
                    text=f"✅ {eyespot_count}/24 œufs - Dates: {dates_summary}",
                    fg="green"
                )
        else:
            self.label_eyespot_indicator.config(text="⚠️ Aucune date Eyespot", fg="orange")

        # Nettoyage des incohérences et doublons
        inconsistencies_fixed = self.clean_data_inconsistencies()
        duplicates_removed = self.clean_photo_duplicates()

        # Mise à jour tracking photos
        self.update_photo_tracking()
        self.update_status()
        self.check_photo_death_conflicts()

        if duplicates_removed > 0 or inconsistencies_fixed > 0:
            # Auto-sauvegarde vers Excel en arrière-plan (ne bloque pas l'interface)
            self._save_excel_background()
        else:
            self.clear_unsaved_indicator()

    def _update_button_color(self, pos, data):
        """Met à jour la couleur d'un bouton"""
        btn = self.buttons[pos]

        if not data['alive']:
            _, bg_color = self.get_death_emoji_and_color(data['death_type'])
            btn.config(bg=bg_color, relief=tk.RAISED, bd=3, highlightthickness=0)
        elif data['hatching_date']:
            today = datetime.now()
            today_str = today.strftime("%d/%m/%Y")
            hatching_str = data['hatching_date']

            if hatching_str == today_str:
                btn.config(bg=COLOR_HATCHED_TODAY, relief=tk.RIDGE, bd=5,
                          highlightthickness=2, highlightbackground='#FFD700')
            else:
                try:
                    hatching_dt = datetime.strptime(hatching_str, "%d/%m/%Y")
                    days = (today - hatching_dt).days
                    if days > 14:
                        btn.config(bg=COLOR_HATCHED_14DAYS, relief=tk.RAISED, bd=3, highlightthickness=0)
                    else:
                        btn.config(bg=COLOR_HATCHED, relief=tk.RAISED, bd=3, highlightthickness=0)
                except:
                    btn.config(bg=COLOR_HATCHED, relief=tk.RAISED, bd=3, highlightthickness=0)
        else:
            btn.config(bg=COLOR_ALIVE, relief=tk.RAISED, bd=3, highlightthickness=0)

    def get_death_emoji_and_color(self, death_type):
        """Retourne emoji et couleur selon type de mort"""
        if death_type == "Dead eyed":
            return "👁️", COLOR_DEAD_EYED
        elif death_type == "Dead larvae":
            return "🐟", COLOR_DEAD_LARVAE
        elif death_type == "Runaway":
            return "🏃", COLOR_RUNAWAY
        elif death_type == "Out of Studies":
            return "📤", COLOR_OUT_OF_STUDIES
        else:
            return "☠️", COLOR_DEAD

    def mark_unsaved(self):
        """Marque qu'il y a des modifications non sauvegardées"""
        self.has_unsaved_changes = True
        self.label_unsaved.config(text="⚠️ Non sauvegardé", fg="red")

    def mark_saved(self):
        """Marque que tout est sauvegardé"""
        self.has_unsaved_changes = False
        self.label_unsaved.config(text="✅ Sauvegardé", fg="green")

    def clear_unsaved_indicator(self):
        """Efface l'indicateur de sauvegarde"""
        self.has_unsaved_changes = False
        self.label_unsaved.config(text="")

    def _save_excel_background(self):
        """Sauvegarde vers Excel en arrière-plan (thread séparé)"""
        # Copier les données nécessaires pour le thread
        plate_num = self.current_plate_number
        cells_data = {pos: dict(data) for pos, data in self.cells_status.items()}
        female_type = self.current_female_type
        fert_date = self.current_fert_date

        def save_worker():
            try:
                plate_name = f"Plaque_{plate_num:03d}.xlsx"
                plate_path = os.path.join(MODIFIED_FOLDER, plate_name)

                if not os.path.exists(plate_path):
                    src_path = os.path.join(PLATES_FOLDER, plate_name)
                    if os.path.exists(src_path):
                        shutil.copy(src_path, plate_path)

                wb = openpyxl.load_workbook(plate_path)
                ws_suivi = wb["Suivi"]

                # Mettre à jour TOUTES les colonnes (statut, mort, photos)
                for row_idx, row in enumerate(ROWS):
                    for col_idx, col in enumerate(COLUMNS):
                        pos = f"{row}{col}"
                        data = cells_data.get(pos, {})
                        suivi_row = row_idx * len(COLUMNS) + col_idx + 2

                        # Statut, death_date, death_type
                        is_alive = data.get('alive', True)
                        ws_suivi.cell(suivi_row, 9).value = "Alive" if is_alive else "Dead"
                        ws_suivi.cell(suivi_row, 10).value = data.get('death_date', '')
                        ws_suivi.cell(suivi_row, 11).value = data.get('death_type', '')

                        # Photos
                        photo_dates = data.get('photo_dates', [])
                        # Effacer les anciennes dates
                        for i in range(20):
                            ws_suivi.cell(suivi_row, 13 + i).value = None
                        # Écrire les nouvelles dates (sans doublons)
                        for i, photo_date in enumerate(photo_dates[:20]):
                            ws_suivi.cell(suivi_row, 13 + i).value = photo_date

                wb.save(plate_path)
                wb.close()

                # Mettre à jour le cache
                self.cache.save_plate_to_cache(plate_num, female_type, fert_date, cells_data)

                print(f"🧹 Auto-correction doublons → Excel sauvegardé (Plaque {plate_num:03d})")

                # Mettre à jour l'indicateur dans le thread principal
                self.root.after(0, self.clear_unsaved_indicator)

            except Exception as e:
                print(f"⚠️ Erreur auto-sauvegarde: {e}")

        thread = threading.Thread(target=save_worker, daemon=True)
        thread.start()

    def clean_data_inconsistencies(self):
        """Nettoie les incohérences de données (ex: alive=True mais death_type non vide)"""
        if not self.cells_status:
            return 0

        fixes = 0
        for pos, data in self.cells_status.items():
            # Si vivant, effacer death_date et death_type
            if data.get('alive', True):
                if data.get('death_date') or data.get('death_type'):
                    data['death_date'] = ""
                    data['death_type'] = ""
                    fixes += 1

        if fixes > 0:
            print(f"🔧 Correction: {fixes} incohérence(s) de données corrigée(s)")

        return fixes

    def clean_photo_duplicates(self):
        """Supprime les doublons de dates de photos pour toutes les cellules"""
        if not self.cells_status or not self.current_plate_number:
            return 0

        total_removed = 0
        for pos, data in self.cells_status.items():
            photo_dates = data.get('photo_dates', [])
            if photo_dates:
                # Garder uniquement les dates uniques (en préservant l'ordre)
                seen = set()
                unique_dates = []
                for date in photo_dates:
                    if date not in seen:
                        seen.add(date)
                        unique_dates.append(date)

                removed = len(photo_dates) - len(unique_dates)
                if removed > 0:
                    total_removed += removed
                    data['photo_dates'] = unique_dates
                    # Recalculer la dernière date photo
                    if unique_dates:
                        data['last_photo_date'] = unique_dates[-1]
                    else:
                        data['last_photo_date'] = ''

                    # Mettre à jour le cache SQLite
                    self.cache.update_cell(
                        self.current_plate_number, pos,
                        photo_dates=unique_dates,
                        last_photo_date=data['last_photo_date']
                    )

        if total_removed > 0:
            print(f"🧹 Nettoyage: {total_removed} doublon(s) de photo supprimé(s)")

        return total_removed

    def toggle_cell(self, pos):
        """Toggle l'état d'une cellule"""
        if pos not in self.cells_status:
            return

        if self.hatching_mode:
            self.handle_hatching_mode(pos)
            return

        if self.photo_mode:
            self.handle_photo_mode(pos)
            return

        # Mode normal : toggle mort/vivant
        data = self.cells_status[pos]

        if data['alive']:
            # Marquer comme mort
            data['alive'] = False
            data['death_date'] = datetime.now().strftime("%d/%m/%Y")
            data['death_type'] = self.death_type_var.get()
            self.selected_dead.add(pos)
            self.newly_marked_dead.add(pos)
        else:
            # Remettre vivant
            data['alive'] = True
            data['death_date'] = ""
            data['death_type'] = ""  # Effacer le type de mort
            self.selected_dead.discard(pos)
            self.newly_marked_dead.discard(pos)

        # Mettre à jour l'affichage
        self._refresh_button(pos)
        self.update_status()
        self.check_photo_death_conflicts()
        self.mark_unsaved()

    def _refresh_button(self, pos):
        """Rafraîchit l'affichage d'un bouton"""
        data = self.cells_status[pos]
        btn = self.buttons[pos]

        cross = data.get('cross', '')
        if MALE_11_15_PATTERN.search(cross):
            btn_text = f"⭕ {cross}" if len(cross) <= 12 else f"⭕ {cross[:10]}..."
        else:
            btn_text = cross if len(cross) <= 15 else cross[:12] + "..."

        indicators = []
        if not data['alive'] and data['death_date']:
            emoji, _ = self.get_death_emoji_and_color(data['death_type'])
            indicators.append(f"{emoji} {data['death_date']}")

        if data['hatching_date']:
            indicators.append(f"🐣 {data['hatching_date']}")

        if indicators:
            btn_text = f"{btn_text}\n" + "\n".join(indicators)

        btn.config(text=btn_text)
        self._update_button_color(pos, data)

    def handle_hatching_mode(self, pos):
        """Gère le mode éclosion"""
        data = self.cells_status[pos]
        modified = False

        if data['hatching_date']:
            if messagebox.askyesno("Annuler éclosion",
                f"L'œuf {pos} est marqué éclos le {data['hatching_date']}.\nAnnuler ?"):
                data['hatching_date'] = ""
                modified = True
        else:
            hatching_date = self.entry_hatching.get().strip()
            if hatching_date:
                data['hatching_date'] = hatching_date
                modified = True

        self._refresh_button(pos)
        self.update_photo_tracking()
        if modified:
            self.mark_unsaved()

    def handle_photo_mode(self, pos):
        """Enregistre une date de photo"""
        data = self.cells_status[pos]

        if not data['alive'] or not data['hatching_date']:
            return

        today_str = datetime.now().strftime("%d/%m/%Y")

        if 'photo_dates' not in data:
            data['photo_dates'] = []

        # Vérifier si la date d'aujourd'hui est déjà enregistrée
        if today_str in data['photo_dates']:
            messagebox.showinfo("Photo déjà enregistrée",
                f"Une photo pour {pos} a déjà été enregistrée aujourd'hui ({today_str}).\n"
                f"Total photos: {len(data['photo_dates'])}")
            return

        data['photo_dates'].append(today_str)
        data['last_photo_date'] = today_str

        # Mettre à jour le cache
        self.cache.update_cell(
            self.current_plate_number, pos,
            photo_dates=data['photo_dates'],
            last_photo_date=today_str
        )

        self.update_photo_tracking()
        self.mark_unsaved()
        messagebox.showinfo("Photo enregistrée", f"Photo pour {pos} enregistrée : {today_str}")

    def show_context_menu(self, event, pos):
        """Affiche un menu contextuel clic droit avec infos et gestion des photos"""
        data = self.cells_status.get(pos)
        if not data:
            return

        menu = tk.Menu(self.root, tearoff=0)

        # Infos de l'alevin
        cross = data.get('cross', 'N/A')
        menu.add_command(label=f"Position : {pos} — {cross}", state=tk.DISABLED)
        menu.add_separator()

        statut = "Vivant" if data['alive'] else "Mort"
        menu.add_command(label=f"Statut : {statut}", state=tk.DISABLED)

        if not data['alive']:
            death_type = data.get('death_type', 'Dead')
            death_date = data.get('death_date', '')
            menu.add_command(label=f"Type de mort : {death_type}", state=tk.DISABLED)
            menu.add_command(label=f"Date de mort : {death_date}", state=tk.DISABLED)
            menu.add_command(
                label="✏️ Modifier la date de mort",
                command=lambda: self.edit_death_date(pos)
            )

        if data.get('hatching_date'):
            menu.add_command(label=f"Date éclosion : {data['hatching_date']}", state=tk.DISABLED)

        # Section photos
        photo_dates = data.get('photo_dates', [])
        if photo_dates:
            menu.add_separator()
            menu.add_command(label=f"📷 Photos ({len(photo_dates)}) :", state=tk.DISABLED)
            for i, pd in enumerate(photo_dates):
                menu.add_command(label=f"   Photo {i+1} : {pd}", state=tk.DISABLED)

            # Alerte si date photo == date mort (hors Out of Studies)
            death_date = data.get('death_date', '')
            death_type = data.get('death_type', '')
            has_conflict = (not data['alive'] and death_date
                           and death_type != "Out of Studies"
                           and death_date in photo_dates)

            if has_conflict:
                menu.add_separator()
                if pos in self.ignored_photo_conflicts:
                    menu.add_command(
                        label="ℹ️ Conflit photo/mort ignoré",
                        state=tk.DISABLED, foreground="gray"
                    )
                    menu.add_command(
                        label="🔔 Rétablir l'alerte",
                        command=lambda: self.unignore_photo_conflict(pos)
                    )
                else:
                    menu.add_command(
                        label="⚠️ ALERTE : Photo le même jour que la mort !",
                        state=tk.DISABLED, foreground="red"
                    )
                    menu.add_command(
                        label="✅ Ignorer cette alerte",
                        command=lambda: self.ignore_photo_conflict(pos)
                    )

            # Option supprimer la dernière photo
            menu.add_separator()
            last_photo = photo_dates[-1]
            menu.add_command(
                label=f"🗑️ Supprimer la dernière photo ({last_photo})",
                command=lambda: self.delete_last_photo(pos)
            )
        else:
            menu.add_separator()
            menu.add_command(label="📷 Aucune photo", state=tk.DISABLED)

        menu.tk_popup(event.x_root, event.y_root)

    def ignore_photo_conflict(self, pos):
        """Ignore l'alerte photo/mort pour une cellule"""
        self.ignored_photo_conflicts.add(pos)
        # Persister dans la DB
        self.cache.add_ignored_conflict(self.current_plate_number, pos)
        self.check_photo_death_conflicts()

    def unignore_photo_conflict(self, pos):
        """Rétablit l'alerte photo/mort pour une cellule"""
        self.ignored_photo_conflicts.discard(pos)
        # Retirer de la DB
        self.cache.remove_ignored_conflict(self.current_plate_number, pos)
        self.check_photo_death_conflicts()

    def edit_death_date(self, pos):
        """Permet de modifier la date de mort d'un alevin via une popup"""
        data = self.cells_status.get(pos)
        if not data or data['alive']:
            return

        # Créer la popup
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Modifier date de mort - {pos}")
        dialog.geometry("350x180")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        # Centrer
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # Label info
        tk.Label(
            dialog,
            text=f"Position : {pos}\nType de mort : {data.get('death_type', 'Dead')}",
            font=("Arial", 10), pady=10
        ).pack()

        # Frame pour la date
        frame_date = tk.Frame(dialog)
        frame_date.pack(pady=10)

        tk.Label(frame_date, text="Nouvelle date (JJ/MM/AAAA) :", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        entry_date = tk.Entry(frame_date, width=12, font=("Arial", 10))
        entry_date.pack(side=tk.LEFT, padx=5)
        entry_date.insert(0, data.get('death_date', ''))
        entry_date.focus_set()
        entry_date.select_range(0, tk.END)

        # Boutons
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=15)

        def validate_and_save():
            new_date = entry_date.get().strip()

            # Vérifier le format
            try:
                datetime.strptime(new_date, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Erreur", "Format de date invalide.\nUtilisez JJ/MM/AAAA (ex: 15/02/2026)")
                return

            # Sauvegarder
            data['death_date'] = new_date

            # Mettre à jour le cache
            self.cache.update_cell(
                self.current_plate_number, pos,
                death_date=new_date
            )

            self._refresh_button(pos)
            self.check_photo_death_conflicts()
            self.mark_unsaved()
            dialog.destroy()
            messagebox.showinfo("Succès", f"Date de mort modifiée pour {pos} :\n{new_date}")

        tk.Button(
            button_frame, text="✅ Valider", command=validate_and_save,
            bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), padx=15
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            button_frame, text="❌ Annuler", command=dialog.destroy,
            bg="#f44336", fg="white", font=("Arial", 10), padx=15
        ).pack(side=tk.LEFT, padx=5)

        # Raccourcis
        dialog.bind('<Return>', lambda e: validate_and_save())
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def delete_last_photo(self, pos):
        """Supprime la dernière date de photo d'un alevin"""
        data = self.cells_status.get(pos)
        if not data:
            return

        photo_dates = data.get('photo_dates', [])
        if not photo_dates:
            return

        removed = photo_dates.pop()

        # Recalculer last_photo_date
        data['last_photo_date'] = photo_dates[-1] if photo_dates else ''

        # Mettre à jour le cache SQLite
        self.cache.update_cell(
            self.current_plate_number, pos,
            photo_dates=data['photo_dates'],
            last_photo_date=data['last_photo_date']
        )

        self._refresh_button(pos)
        self.update_photo_tracking()
        self.check_photo_death_conflicts()
        self.mark_unsaved()
        messagebox.showinfo("Photo supprimée", f"Photo du {removed} supprimée pour {pos}")

    def toggle_hatching_mode(self):
        """Active/désactive le mode éclosion"""
        self.hatching_mode = not self.hatching_mode
        if self.hatching_mode:
            self.btn_hatching.config(text="🐣 Mode éclosion ACTIF (O)")
            self.photo_mode = False
            self.btn_photo_mode.config(text="📷 Mode photo (P)", bg="#FF9800")
        else:
            self.btn_hatching.config(text="🐣 Mode éclosion (O)")

    def toggle_photo_mode(self):
        """Active/désactive le mode photo"""
        self.photo_mode = not self.photo_mode
        if self.photo_mode:
            self.btn_photo_mode.config(text="📷 Mode photo ACTIF", bg="#4CAF50")
            self.hatching_mode = False
            self.btn_hatching.config(text="🐣 Mode éclosion")
        else:
            self.btn_photo_mode.config(text="📷 Mode photo (P)", bg="#FF9800")

    def apply_eyespot(self):
        """Applique la date eyespot aux cellules vivantes sans date"""
        if not self.current_plate_number:
            return

        date = self.entry_eyespot.get().strip()
        if not date:
            return

        count = 0
        for pos, data in self.cells_status.items():
            if data['alive'] and not data.get('eyespot_date'):
                data['eyespot_date'] = date
                count += 1

        if count > 0:
            self.mark_unsaved()
        messagebox.showinfo("Eyespot", f"Date appliquée à {count} cellules")

    def apply_out_of_studies(self):
        """Applique 'Out of Studies' à tous les alevins éclos depuis plus de 14 jours"""
        if not self.current_plate_number:
            messagebox.showwarning("Attention", "Chargez d'abord une plaque")
            return

        today = datetime.now()
        today_str = today.strftime("%d/%m/%Y")
        count = 0

        for pos, data in self.cells_status.items():
            # Seulement les alevins vivants avec une date d'éclosion
            if data['alive'] and data.get('hatching_date'):
                try:
                    hatching_dt = datetime.strptime(data['hatching_date'], "%d/%m/%Y")
                    days_since_hatching = (today - hatching_dt).days

                    if days_since_hatching > 14:
                        # Marquer comme "Out of Studies"
                        data['alive'] = False
                        data['death_date'] = today_str
                        data['death_type'] = "Out of Studies"
                        self.selected_dead.add(pos)
                        self.newly_marked_dead.add(pos)
                        self._refresh_button(pos)
                        count += 1
                except:
                    pass

        if count > 0:
            self.mark_unsaved()
            self.update_status()
            messagebox.showinfo("Out of Studies",
                f"📤 {count} alevin(s) marqué(s) 'Out of Studies'\n"
                f"(éclos depuis plus de 14 jours)\n\n")
        else:
            messagebox.showinfo("Out of Studies",
                "Aucun alevin éligible trouvé\n"
                "(aucun alevin vivant éclos depuis plus de 14 jours)")

    def update_photo_tracking(self):
        """Met à jour le tracking visuel des photos"""
        if not self.current_plate_number:
            self.blink_cells.clear()
            return

        old_blink = self.blink_cells.copy()
        self.blink_cells.clear()

        today = datetime.now()

        for pos, data in self.cells_status.items():
            if data['alive'] and data.get('hatching_date'):
                needs_photo = False

                if not data.get('last_photo_date'):
                    try:
                        hatching_dt = datetime.strptime(data['hatching_date'], "%d/%m/%Y")
                        if (today - hatching_dt).days >= 3:
                            needs_photo = True
                    except:
                        pass
                else:
                    try:
                        last_photo_dt = datetime.strptime(data['last_photo_date'], "%d/%m/%Y")
                        if (today - last_photo_dt).days >= 3:
                            needs_photo = True
                    except:
                        pass

                btn = self.buttons[pos]
                current_text = btn.cget('text')

                if needs_photo:
                    self.blink_cells.add(pos)
                    if not current_text.startswith('📸'):
                        btn.config(text=f"📸📸 {current_text}")
                elif current_text.startswith('📸📸 '):
                    btn.config(text=current_text[4:])

        # Indicateur
        count = len(self.blink_cells)
        if count > 0:
            self.label_photo_indicator.config(
                text=f"📸 {count} alevin(s) à photographier",
                fg="red"
            )
        else:
            self.label_photo_indicator.config(text="", fg="black")

        # Prochaine plaque
        self.find_next_plate_to_photograph()

    def find_next_plate_to_photograph(self):
        """Trouve la prochaine plaque avec photos à prendre"""
        if not self.current_plate_number:
            return

        plates = self.cache.get_plates_needing_photos()

        # Chercher la prochaine après la plaque actuelle
        for plate_num in plates:
            if plate_num > self.current_plate_number:
                self.label_next_photo.config(text=f"📷 Prochaine: {plate_num:03d}")
                return

        self.label_next_photo.config(text="")

    def update_status(self):
        """Met à jour la barre de statut"""
        if not self.current_plate_number:
            self.label_status.config(text="Aucune plaque chargée")
            return

        nb_dead = len(self.selected_dead)
        nb_alive = 24 - nb_dead
        nb_new = len(self.newly_marked_dead)

        text = f"Plaque {self.current_plate_number:03d} | ❌ Morts: {nb_dead} | ✅ Vivants: {nb_alive}"
        if nb_new > 0:
            text += f" | 🆕 Nouveaux: {nb_new}"

        self.label_status.config(text=text)

    def check_photo_death_conflicts(self):
        """Vérifie si des alevins morts (hors Out of Studies) ont une photo le même jour que leur mort"""
        if not self.current_plate_number:
            self.label_photo_death_warning.config(text="")
            return

        conflicts = []
        for pos, data in self.cells_status.items():
            if pos in self.ignored_photo_conflicts:
                continue
            if not data['alive'] and data.get('death_date'):
                death_type = data.get('death_type', '')
                if death_type == "Out of Studies":
                    continue
                death_date = data['death_date']
                photo_dates = data.get('photo_dates', [])
                if death_date in photo_dates:
                    conflicts.append(pos)

        if conflicts:
            positions = ", ".join(sorted(conflicts))
            self.label_photo_death_warning.config(
                text=f"⚠️ Photo = jour de mort : {positions} (clic droit pour corriger)"
            )
        else:
            self.label_photo_death_warning.config(text="")

    def create_tooltip(self, widget, pos):
        """Tooltip au survol - affiche infos non visibles sur le bouton avec scroll pour les photos"""
        def show(event):
            if not self.current_plate_number or pos not in self.cells_status:
                return

            data = self.cells_status[pos]
            has_content = False

            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+15}+{event.y_root+10}")

            # Frame principal avec fond jaune
            main_frame = tk.Frame(tooltip, background="#FFFFD0", relief=tk.SOLID, borderwidth=1)
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Type de mort (la date est déjà visible)
            if not data['alive']:
                death_type = data.get('death_type', 'Dead')
                emoji, _ = self.get_death_emoji_and_color(death_type)
                tk.Label(
                    main_frame, text=f"{emoji} Type: {death_type}",
                    background="#FFFFD0", font=("Arial", 9), padx=8, pady=2, justify=tk.LEFT
                ).pack(anchor='w')
                has_content = True

            # Photos (info non visible sur le bouton)
            if data['alive'] and data.get('hatching_date'):
                photo_dates = data.get('photo_dates', [])
                if photo_dates:
                    # Titre
                    tk.Label(
                        main_frame, text=f"📷 Photos: {len(photo_dates)}",
                        background="#FFFFD0", font=("Arial", 9, "bold"), padx=8, pady=2
                    ).pack(anchor='w')

                    # Frame avec scrollbar pour les photos
                    if len(photo_dates) > 8:
                        # Si plus de 8 photos, ajouter une scrollbar
                        scroll_frame = tk.Frame(main_frame, background="#FFFFD0")
                        scroll_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=2)

                        # Canvas et Scrollbar
                        canvas = tk.Canvas(scroll_frame, background="#FFFFD0",
                                         height=160, width=200, highlightthickness=0)
                        scrollbar = tk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
                        scrollable_frame = tk.Frame(canvas, background="#FFFFD0")

                        scrollable_frame.bind(
                            "<Configure>",
                            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                        )

                        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                        canvas.configure(yscrollcommand=scrollbar.set)

                        # Afficher toutes les photos dans le frame scrollable
                        for i, pd in enumerate(photo_dates):
                            tk.Label(
                                scrollable_frame, text=f"   Photo {i+1}: {pd}",
                                background="#FFFFD0", font=("Arial", 9), justify=tk.LEFT
                            ).pack(anchor='w')

                        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

                        # Positionner le scroll en bas pour voir la dernière photo
                        tooltip.update_idletasks()
                        canvas.yview_moveto(1.0)
                    else:
                        # Si 8 photos ou moins, affichage simple sans scroll
                        for i, pd in enumerate(photo_dates):
                            tk.Label(
                                main_frame, text=f"   Photo {i+1}: {pd}",
                                background="#FFFFD0", font=("Arial", 9), padx=8, pady=1, justify=tk.LEFT
                            ).pack(anchor='w')

                    has_content = True
                else:
                    tk.Label(
                        main_frame, text=f"📷 Aucune photo",
                        background="#FFFFD0", font=("Arial", 9), padx=8, pady=2
                    ).pack(anchor='w')
                    has_content = True

            if not has_content:
                tooltip.destroy()
                return

            widget.tooltip = tooltip

        def hide(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip

        widget.bind('<Enter>', show)
        widget.bind('<Leave>', hide)

    def apply_photo_to_all_alive(self):
        """Applique la date de photo à tous les alevins vivants éclos"""
        if not self.current_plate_number:
            messagebox.showwarning("Attention", "Chargez d'abord une plaque")
            return

        today_str = datetime.now().strftime("%d/%m/%Y")
        count = 0

        for pos, data in self.cells_status.items():
            # Alevin vivant ET éclos
            if data['alive'] and data.get('hatching_date'):
                if 'photo_dates' not in data:
                    data['photo_dates'] = []

                # Ajouter si pas déjà présent
                if today_str not in data['photo_dates']:
                    data['photo_dates'].append(today_str)
                    data['last_photo_date'] = today_str

                    # Mettre à jour le cache
                    self.cache.update_cell(
                        self.current_plate_number, pos,
                        photo_dates=data['photo_dates'],
                        last_photo_date=today_str
                    )
                    count += 1

        self.update_photo_tracking()
        if count > 0:
            self.mark_unsaved()
        messagebox.showinfo("Photos", f"Date photo appliquée à {count} alevins vivants éclos")

    def plate_has_photos_needed(self, plate_num):
        """Vérifie si une plaque a des alevins nécessitant une photo"""
        cached = self.cache.get_plate_data(plate_num)
        if not cached:
            return False

        today = datetime.now()

        for pos, data in cached['cells'].items():
            if not data.get('alive', True) or not data.get('hatching_date'):
                continue

            # Vérifier si photo nécessaire
            last_photo = data.get('last_photo_date', '')
            hatching = data.get('hatching_date', '')

            try:
                if not last_photo:
                    hatching_dt = datetime.strptime(hatching, "%d/%m/%Y")
                    if (today - hatching_dt).days >= 3:
                        return True
                else:
                    last_dt = datetime.strptime(last_photo, "%d/%m/%Y")
                    if (today - last_dt).days >= 3:
                        return True
            except:
                pass

        return False

    def load_previous_plate_silent(self):
        """Charge la plaque précédente SANS boîte de dialogue"""
        if self.current_plate_number is None:
            return

        prev_num = self.current_plate_number - 1
        if prev_num < 1:
            return

        self.entry_plate.delete(0, tk.END)
        self.entry_plate.insert(0, str(prev_num))
        self.load_plate_silent()

    def load_next_plate_silent(self):
        """Charge la plaque suivante SANS boîte de dialogue"""
        if self.current_plate_number is None:
            return

        next_num = self.current_plate_number + 1
        self.entry_plate.delete(0, tk.END)
        self.entry_plate.insert(0, str(next_num))
        self.load_plate_silent()

    def load_plate_silent(self):
        """Charge une plaque SANS afficher de message ni dialogue de sauvegarde"""
        plate_str = self.entry_plate.get().strip()
        if not plate_str:
            return

        try:
            plate_num = int(plate_str)
        except ValueError:
            return

        # Essayer le cache d'abord
        cached_data = self.cache.get_plate_data(plate_num)

        if cached_data:
            self._load_from_cache(plate_num, cached_data)
        else:
            # Synchroniser puis charger
            if ExcelSyncer.sync_plate_to_cache(plate_num, self.cache):
                cached_data = self.cache.get_plate_data(plate_num)
                if cached_data:
                    self._load_from_cache(plate_num, cached_data)

    def load_previous_plate(self):
        """Charge la plaque précédente"""
        if self.current_plate_number is None:
            messagebox.showwarning("Attention", "Chargez d'abord une plaque")
            return

        if self.current_plate_number <= 1:
            messagebox.showinfo("Info", "C'est déjà la première plaque")
            return

        self.entry_plate.delete(0, tk.END)
        self.entry_plate.insert(0, str(self.current_plate_number - 1))
        self.load_plate()

    def load_next_plate(self):
        """Charge la plaque suivante"""
        if self.current_plate_number is None:
            messagebox.showwarning("Attention", "Chargez d'abord une plaque")
            return

        self.entry_plate.delete(0, tk.END)
        self.entry_plate.insert(0, str(self.current_plate_number + 1))
        self.load_plate()

    def _show_save_dialog(self):
        """Affiche dialogue de sauvegarde personnalisé"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Vérification")
        dialog.geometry("450x200")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        # Centrer
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        user_choice = {'action': None}

        # Message
        tk.Label(
            dialog,
            text=f"Une plaque est déjà chargée ({self.current_plate}).\nQue souhaitez-vous faire ?",
            font=("Arial", 10), pady=20
        ).pack()

        # Boutons
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=10)

        def on_save():
            user_choice['action'] = 'save'
            dialog.destroy()

        def on_discard():
            user_choice['action'] = 'discard'
            dialog.destroy()

        def on_cancel():
            user_choice['action'] = 'cancel'
            dialog.destroy()

        save_btn = tk.Button(button_frame, text="💾 Sauvegarder", command=on_save,
                  bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), padx=10)
        save_btn.pack(side=tk.LEFT, padx=5)
        save_btn.focus_set()  # Focus sur le bouton sauvegarder

        tk.Button(button_frame, text="🗑️ Ignorer", command=on_discard,
                  bg="#FF9800", fg="white", font=("Arial", 10), padx=10).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="❌ Annuler", command=on_cancel,
                  bg="#f44336", fg="white", font=("Arial", 10), padx=10).pack(side=tk.LEFT, padx=5)

        # Raccourcis clavier
        dialog.bind('<Return>', lambda e: on_save())
        dialog.bind('<space>', lambda e: on_save())

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        self.root.wait_window(dialog)

        return user_choice['action']

    def _save_plate_silent(self):
        """Sauvegarde silencieuse (sans dialogue de confirmation)"""
        if not self.current_plate_number or not self.cells_status:
            return False

        try:
            plate_name = f"Plaque_{self.current_plate_number:03d}.xlsx"
            plate_path = os.path.join(MODIFIED_FOLDER, plate_name)

            if not os.path.exists(plate_path):
                src_path = os.path.join(PLATES_FOLDER, plate_name)
                if os.path.exists(src_path):
                    shutil.copy(src_path, plate_path)

            wb = openpyxl.load_workbook(plate_path)
            ws_disp = wb["Disposition"]
            ws_suivi = wb["Suivi"]

            # Couleurs Excel
            fills = {
                "dead": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                "eyed": PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid"),
                "larvae": PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"),
                "alive": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
                "out_of_studies": PatternFill(start_color="505050", end_color="505050", fill_type="solid")
            }

            for row_idx, row in enumerate(ROWS):
                for col_idx, col in enumerate(COLUMNS):
                    pos = f"{row}{col}"
                    data = self.cells_status.get(pos, {})
                    suivi_row = row_idx * len(COLUMNS) + col_idx + 2

                    is_alive = data.get('alive', True)

                    # Couleur dans Disposition
                    cell_disp = ws_disp.cell(row_idx + 2, col_idx + 2)
                    if is_alive:
                        cell_disp.fill = fills["alive"]
                    else:
                        death_type = data.get('death_type', 'Dead')
                        if death_type == "Dead eyed":
                            cell_disp.fill = fills["eyed"]
                        elif death_type == "Dead larvae":
                            cell_disp.fill = fills["larvae"]
                        elif death_type == "Out of Studies":
                            cell_disp.fill = fills["out_of_studies"]
                        else:
                            cell_disp.fill = fills["dead"]

                    # Suivi
                    ws_suivi.cell(suivi_row, 9).value = "Alive" if is_alive else "Dead"
                    ws_suivi.cell(suivi_row, 10).value = data.get('death_date', '')
                    ws_suivi.cell(suivi_row, 11).value = data.get('death_type', '')
                    ws_suivi.cell(suivi_row, 8).value = data.get('hatching_date', '')
                    ws_suivi.cell(suivi_row, 7).value = data.get('eyespot_date', '')

                    # Photos
                    photo_dates = data.get('photo_dates', [])
                    for i, photo_date in enumerate(photo_dates[:20]):
                        ws_suivi.cell(suivi_row, 13 + i).value = photo_date

            wb.save(plate_path)
            wb.close()

            # Mettre à jour le cache
            self.cache.save_plate_to_cache(
                self.current_plate_number,
                self.current_female_type,
                self.current_fert_date,
                self.cells_status
            )

            self.newly_marked_dead.clear()
            self.mark_saved()
            print(f"✅ Sauvegarde: Plaque {self.current_plate_number:03d}")
            return True

        except Exception as e:
            print(f"⚠️ Erreur sauvegarde: {e}")
            return False

    def resync_current_plate(self):
        """Resynchronise la plaque actuelle depuis Excel"""
        if not self.current_plate_number:
            return

        if ExcelSyncer.sync_plate_to_cache(self.current_plate_number, self.cache):
            self.load_plate()
            messagebox.showinfo("Resync", "Plaque resynchronisée depuis Excel")

    def save_plate(self):
        """Sauvegarde la plaque avec toutes les modifications (comme LIVE)"""
        if not self.current_plate_number:
            messagebox.showwarning("Attention", "Aucune plaque chargée")
            return

        if not messagebox.askyesno("Confirmation", f"Sauvegarder les modifications de {self.current_plate} ?"):
            return

        # Sauvegarder
        if self._save_plate_silent():
            # Compter les morts par date ET par type
            dates_summary = {}
            types_summary = {"Dead": 0, "Dead eyed": 0, "Dead larvae": 0, "Runaway": 0, "Out of Studies": 0}

            for pos in self.selected_dead:
                date = self.cells_status[pos].get('death_date', 'Non datée')
                death_type = self.cells_status[pos].get('death_type', 'Dead')

                dates_summary[date] = dates_summary.get(date, 0) + 1
                types_summary[death_type] = types_summary.get(death_type, 0) + 1

            dates_text = "\n".join([f"  • {date}: {count} œuf(s)" for date, count in sorted(dates_summary.items())])
            types_text = "\n".join([f"  • {type_name}: {count}" for type_name, count in types_summary.items() if count > 0])

            save_path = os.path.join(MODIFIED_FOLDER, self.current_plate)

            messagebox.showinfo("Succès",
                f"✅ Plaque sauvegardée !\n\n"
                f"📂 Emplacement:\n{save_path}\n\n"
                f"📊 Modifications:\n"
                f"  • Total morts: {len(self.selected_dead)}\n"
                f"  • Nouveaux morts: {len(self.newly_marked_dead)}\n"
                f"  • Fertilisation: {self.current_fert_date}\n\n"
                f"📅 Morts par date:\n{dates_text if dates_text else '  Aucun'}\n\n"
                f"🔬 Morts par type:\n{types_text if types_text else '  Aucun'}")

            self.newly_marked_dead.clear()
            self.update_photo_tracking()
            self.check_photo_death_conflicts()
        else:
            messagebox.showerror("Erreur", "Erreur lors de la sauvegarde")

    def _setup_menu_bar(self):
        """Crée la barre de menu de l'application"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # Menu File
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(
            label="Edit Cell Data...",
            command=self.open_complete_editor,
            accelerator="Ctrl+E"
        )
        file_menu.add_separator()
        file_menu.add_command(
            label="Save Plate",
            command=self.save_plate,
            accelerator="Ctrl+S"
        )
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Menu View
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_command(
            label="Refresh Cache",
            command=self.resync_current_plate,
            accelerator="Ctrl+R"
        )

        # Menu Help
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)

    def open_complete_editor(self):
        """Ouvre l'éditeur complet de données de cellule"""
        if not self.current_plate_number:
            messagebox.showwarning("Attention", "Chargez d'abord une plaque")
            return

        # Ouvrir l'éditeur avec la première cellule par défaut
        CompleteCellEditor(self.root, self, initial_pos="A1")

    def show_about(self):
        """Affiche la boîte de dialogue About"""
        about_text = """CORELAC Plate Manager [FAST]
Version 2.0 - Optimized Edition

🧬 Application de gestion des plaques CORELAC
   pour le suivi des alevins de poissons

📊 Fonctionnalités:
  • Cache SQLite pour chargement rapide
  • Synchronisation automatique Excel ↔ SQLite
  • Gestion des statuts (Alive/Dead/Runaway/Out of Studies)
  • Suivi des éclosions et photos
  • Détection des conflits photo/mort

👨‍💻 Développé pour:
   INRAE_USMB - UMR CARRTEL
    Projet CORELAC - Étude des résistances au réchauffement chez les poissons

✍️ Auteur:
   Quentin Godeaux

🛠️ Technologies:
   Python 3.12 • Tkinter • OpenPyXL • SQLite

📅 Dernière mise à jour: Mars 2026

📜 Licence: GNU General Public License v3.0
   Ce logiciel est un logiciel libre; vous pouvez le redistribuer
   et/ou le modifier selon les termes de la Licence Publique Générale GNU.
        """

        messagebox.showinfo("À propos", about_text)


# ============ MAIN ============
if __name__ == "__main__":
    root = tk.Tk()
    app = PlateManagerFast(root)
    root.mainloop()
