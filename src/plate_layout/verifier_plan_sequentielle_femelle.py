"""
Verifplan_femelles_groupees.py - VÉRIFICATION AVEC REGROUPEMENT PAR FEMELLE

But :
- Vérifier la répartition correcte des 300 croisements
- Vérifier le regroupement par femelle (croisements B×B + L×B adjacents)
- Chaque croisement : 16 œufs → 4 plaques (8 à 5°C + 8 à 9°C)
- 200 plaques totales
- Analyse du regroupement des femelles par plaque
- Lecture directe depuis les fichiers Excel (pas de JSON requis)
"""

import os
import csv
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from collections import defaultdict
import openpyxl

# ---------------- CONFIG ----------------
input_csv = r"C:\IE\Etudes\ET_Corégone\ET_Corélac\CORELAC_300_plaques_Aléa\12_grilles_5x5_CORELAC_LB-MATRICE.csv"
plaques_dir = r"C:\IE\Etudes\ET_Corélac\CORELAC_300_plaques_24_femelles_groupées"
output_dir = plaques_dir  # Les rapports seront dans le même dossier

if not os.path.exists(plaques_dir):
    print(f"❌ ERREUR : Dossier non trouvé : {plaques_dir}")
    exit(1)

# Configuration
nb_plaques_total = 200
lignes = ['A', 'B', 'C', 'D']
colonnes = [1, 2, 3, 4, 5, 6]
capacity_per_plaque = 24

print("="*70)
print("🔍 VÉRIFICATION DES PLAQUES - REGROUPEMENT PAR FEMELLE")
print("="*70)
print(f"Configuration:")
print(f"  • Nombre de plaques attendu: {nb_plaques_total}")
print(f"  • Positions par plaque: {capacity_per_plaque}")
print(f"  • Dossier: {plaques_dir}")
print("="*70 + "\n")

# --- LECTURE DU CSV SOURCE ---
print("📖 Lecture du fichier CSV source...")
groupes = []
groupes_noms = []

try:
    with open(input_csv, "r", encoding="utf-8") as f:
        lines = f.readlines()
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            if line and '_G' in line and ',' not in line:
                groupe_name = line
                groupe_croisements = []
                i += 1
                
                if i < len(lines):
                    i += 1
                
                for _ in range(5):
                    if i < len(lines):
                        croi_line = lines[i].strip()
                        if croi_line and ',' in croi_line:
                            parts = croi_line.split(',')
                            for croi in parts[1:6]:
                                croi = croi.strip()
                                if croi and 'x' in croi:
                                    groupe_croisements.append(croi)
                        i += 1
                
                if len(groupe_croisements) == 25:
                    groupes.append(groupe_croisements)
                    groupes_noms.append(groupe_name)
            else:
                i += 1

except FileNotFoundError:
    print(f"❌ ERREUR : Fichier non trouvé : {input_csv}")
    exit(1)

total_croisements = sum(len(g) for g in groupes)
print(f"✅ {len(groupes)} groupes extraits")
print(f"✅ {total_croisements} croisements au total")

all_croisements = [croi for groupe in groupes for croi in groupe]

# --- LECTURE DES PLAQUES EXCEL ---
print("\n📂 Lecture des plaques Excel...")

plaques_data = {}
excel_files = [f for f in os.listdir(plaques_dir) if f.startswith("Plaque_") and f.endswith(".xlsx")]

if not excel_files:
    print(f"❌ ERREUR : Aucun fichier Excel trouvé dans {plaques_dir}")
    exit(1)

print(f"  📊 {len(excel_files)} fichiers Excel trouvés")
print(f"  ⏳ Lecture en cours...")

for idx, excel_file in enumerate(sorted(excel_files), 1):
    if idx % 20 == 0:
        print(f"     ... {idx}/{len(excel_files)} plaques lues")
    
    plaque_name = excel_file.replace(".xlsx", "")
    file_path = os.path.join(plaques_dir, excel_file)
    
    try:
        # Lire la feuille "Infos" pour la température
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Lire la température depuis la feuille Infos
        info_sheet = wb["Infos"]
        temp = None
        for row in info_sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row[0] and "Temperature" in str(row[0]):
                temp_str = str(row[1])
                if "5" in temp_str:
                    temp = 5
                elif "9" in temp_str:
                    temp = 9
                break
        
        # Si pas trouvé dans Infos, déduire du numéro de plaque
        if temp is None:
            plaque_num = int(plaque_name.split("_")[1])
            temp = 5 if plaque_num <= 100 else 9
        
        # Lire les croisements depuis la feuille Disposition
        disp_sheet = wb["Disposition"]
        wells = {}
        
        for row_idx, row in enumerate(disp_sheet.iter_rows(min_row=2, max_row=5, values_only=True), 0):
            ligne = lignes[row_idx]
            for col_idx, cell_value in enumerate(row[1:7], 0):  # Colonnes 1-6
                col = colonnes[col_idx]
                pos = f"{ligne}{col}"
                if cell_value and str(cell_value).strip() and 'x' in str(cell_value):
                    wells[pos] = str(cell_value).strip()
        
        plaques_data[plaque_name] = {
            'temp': temp,
            'wells': wells
        }
        
        wb.close()
        
    except Exception as e:
        print(f"  ⚠️ Erreur lors de la lecture de {excel_file}: {e}")
        continue

print(f"  ✅ {len(plaques_data)} plaques chargées avec succès\n")

if len(plaques_data) == 0:
    print("❌ ERREUR : Aucune plaque n'a pu être lue")
    exit(1)

# --- VÉRIFICATIONS GLOBALES ---
print("🔍 Vérifications globales...")

# 1. Nombre de plaques
if len(plaques_data) != nb_plaques_total:
    print(f"  ⚠️ {len(plaques_data)} plaques trouvées (attendu: {nb_plaques_total})")
else:
    print(f"  ✅ {nb_plaques_total} plaques présentes")

# 2. Compter les occurrences par croisement
occ_by_crois = {c: {"5°C": 0, "9°C": 0, "Total": 0} for c in all_croisements}

for plaque_name, data in plaques_data.items():
    temp_key = "5°C" if data['temp'] == 5 else "9°C"
    for croi in data['wells'].values():
        if croi in occ_by_crois:
            occ_by_crois[croi][temp_key] += 1
            occ_by_crois[croi]["Total"] += 1
        else:
            # Croisement non attendu
            if croi not in occ_by_crois:
                occ_by_crois[croi] = {"5°C": 0, "9°C": 0, "Total": 0}
            occ_by_crois[croi][temp_key] += 1
            occ_by_crois[croi]["Total"] += 1

# 3. Vérifier les anomalies de répartition
problemes_repartition = []
for croi, counts in occ_by_crois.items():
    if counts["5°C"] != 8 or counts["9°C"] != 8 or counts["Total"] != 16:
        problemes_repartition.append({
            'croisement': croi,
            '5C': counts["5°C"],
            '9C': counts["9°C"],
            'total': counts["Total"]
        })

if problemes_repartition:
    print(f"  ❌ {len(problemes_repartition)} croisements avec anomalie de répartition")
    for p in problemes_repartition[:5]:  # Afficher les 5 premiers
        print(f"     • {p['croisement']}: {p['5C']}+{p['9C']}={p['total']} (attendu: 8+8=16)")
    if len(problemes_repartition) > 5:
        print(f"     ... et {len(problemes_repartition) - 5} autres")
else:
    print(f"  ✅ Tous les croisements correctement répartis (8+8=16 œufs)")

# 4. Vérifier le remplissage des plaques
plaques_incompletes = []
for plaque_name, data in plaques_data.items():
    if len(data['wells']) != capacity_per_plaque:
        plaques_incompletes.append({
            'plaque': plaque_name,
            'positions': len(data['wells'])
        })

if plaques_incompletes:
    print(f"  ⚠️ {len(plaques_incompletes)} plaques incomplètes")
    for p in plaques_incompletes[:5]:
        print(f"     • {p['plaque']}: {p['positions']}/24 positions")
    if len(plaques_incompletes) > 5:
        print(f"     ... et {len(plaques_incompletes) - 5} autres")
else:
    print(f"  ✅ Toutes les plaques complètes (24/24 positions)")

# --- ANALYSE DU REGROUPEMENT PAR FEMELLE ---
print("\n👩‍🔬 Analyse du regroupement par femelle...")

# Extraire les femelles de chaque croisement
def extract_female(croisement):
    """Extrait la femelle d'un croisement (après le 'x')"""
    if 'x' in croisement:
        return croisement.split('x')[1]
    return None

# Analyser chaque plaque
plaques_regroupement = {}
for plaque_name, data in plaques_data.items():
    # Compter les croisements par femelle dans cette plaque
    femelles_count = defaultdict(lambda: {'positions': [], 'croisements': []})
    
    for pos, croi in data['wells'].items():
        femelle = extract_female(croi)
        if femelle:
            femelles_count[femelle]['positions'].append(pos)
            femelles_count[femelle]['croisements'].append(croi)
    
    plaques_regroupement[plaque_name] = {
        'temp': data['temp'],
        'femelles': dict(femelles_count),
        'nb_femelles': len(femelles_count)
    }

# Statistiques sur le regroupement
femelles_par_plaque = [info['nb_femelles'] for info in plaques_regroupement.values()]
avg_femelles = sum(femelles_par_plaque) / len(femelles_par_plaque) if femelles_par_plaque else 0

print(f"  📊 Nombre moyen de femelles par plaque: {avg_femelles:.2f}")
print(f"  📊 Min: {min(femelles_par_plaque)} femelles, Max: {max(femelles_par_plaque)} femelles")

# Vérifier l'adjacence des colonnes pour une même femelle
adjacence_ok = 0
adjacence_problemes = []

for plaque_name, info in plaques_regroupement.items():
    for femelle, data_fem in info['femelles'].items():
        positions = data_fem['positions']
        # Extraire les colonnes
        cols = sorted(set([int(pos[1]) for pos in positions]))
        
        # Vérifier si les colonnes sont adjacentes
        if len(cols) > 1:
            is_adjacent = all(cols[i+1] - cols[i] == 1 for i in range(len(cols)-1))
            if is_adjacent:
                adjacence_ok += 1
            else:
                adjacence_problemes.append({
                    'plaque': plaque_name,
                    'femelle': femelle,
                    'colonnes': cols,
                    'croisements': data_fem['croisements']
                })

if adjacence_problemes:
    print(f"  ⚠️ {len(adjacence_problemes)} cas où les colonnes d'une femelle ne sont pas adjacentes")
    for p in adjacence_problemes[:3]:
        print(f"     • {p['plaque']} - {p['femelle']}: colonnes {p['colonnes']}")
    if len(adjacence_problemes) > 3:
        print(f"     ... et {len(adjacence_problemes) - 3} autres")
else:
    print(f"  ✅ Toutes les femelles avec plusieurs colonnes sont bien regroupées en colonnes adjacentes")

# --- ANALYSE PAR TYPE DE CROISEMENT ---
print("\n🧬 Analyse par type de croisement...")

types_croisements = {
    'B×B': [],
    'L×L': [],
    'L×B': [],
    'B×L': []
}

for croi in all_croisements:
    if 'B_M' in croi and 'B_F' in croi:
        types_croisements['B×B'].append(croi)
    elif 'L_M' in croi and 'L_F' in croi:
        types_croisements['L×L'].append(croi)
    elif 'L_M' in croi and 'B_F' in croi:
        types_croisements['L×B'].append(croi)
    elif 'B_M' in croi and 'L_F' in croi:
        types_croisements['B×L'].append(croi)

for type_nom, croisements_type in types_croisements.items():
    print(f"  • {type_nom}: {len(croisements_type)} croisements")

# --- RAPPORTS CSV ---
print("\n📋 Génération des rapports...")

# 1. Rapport croisements
rapport_crois_file = os.path.join(output_dir, "rapport_croisements_femelles.csv")
with open(rapport_crois_file, "w", newline='', encoding='utf-8') as csvf:
    writer = csv.writer(csvf)
    writer.writerow(["Croisement", "Femelle", "Type", "Occ_5°C", "Occ_9°C", "Total", "Attendu", "Statut"])
    
    for c in sorted(occ_by_crois.keys()):
        femelle = extract_female(c)
        # Déterminer le type
        if 'B_M' in c and 'B_F' in c:
            type_croi = "B×B"
        elif 'L_M' in c and 'L_F' in c:
            type_croi = "L×L"
        elif 'L_M' in c and 'B_F' in c:
            type_croi = "L×B"
        elif 'B_M' in c and 'L_F' in c:
            type_croi = "B×L"
        else:
            type_croi = "Inconnu"
        
        a = occ_by_crois[c]["5°C"]
        b = occ_by_crois[c]["9°C"]
        total = occ_by_crois[c]["Total"]
        status = "OK" if (a == 8 and b == 8 and total == 16) else "ERREUR"
        writer.writerow([c, femelle, type_croi, a, b, total, 16, status])

print(f"  ✅ {rapport_crois_file}")

# 2. Rapport plaques avec analyse femelles
rapport_plaques_file = os.path.join(output_dir, "rapport_plaques_femelles.csv")
with open(rapport_plaques_file, "w", newline='', encoding='utf-8') as csvf:
    writer = csv.writer(csvf)
    writer.writerow(["Plaque", "Température", "Nb_positions", "Nb_femelles", "Femelles_présentes", "Statut"])
    
    for plaque_name in sorted(plaques_data.keys(), key=lambda x: int(x.split("_")[1])):
        data = plaques_data[plaque_name]
        info_regr = plaques_regroupement[plaque_name]
        
        temp = f"{data['temp']}°C"
        nb_pos = len(data['wells'])
        nb_fem = info_regr['nb_femelles']
        femelles_list = ', '.join(sorted(info_regr['femelles'].keys()))
        
        status = "OK" if nb_pos == capacity_per_plaque else "ERREUR"
        writer.writerow([plaque_name, temp, nb_pos, nb_fem, femelles_list, status])

print(f"  ✅ {rapport_plaques_file}")

# 3. Rapport détaillé par femelle
rapport_femelles_file = os.path.join(output_dir, "rapport_par_femelle.csv")
with open(rapport_femelles_file, "w", newline='', encoding='utf-8') as csvf:
    writer = csv.writer(csvf)
    writer.writerow(["Femelle", "Nb_croisements_total", "Nb_plaques_5C", "Nb_plaques_9C", "Plaques_5C", "Plaques_9C"])
    
    # Regrouper par femelle
    femelles_data = defaultdict(lambda: {'croisements': set(), 'plaques_5C': set(), 'plaques_9C': set()})
    
    for plaque_name, data in plaques_data.items():
        for croi in data['wells'].values():
            femelle = extract_female(croi)
            if femelle:
                femelles_data[femelle]['croisements'].add(croi)
                if data['temp'] == 5:
                    femelles_data[femelle]['plaques_5C'].add(plaque_name)
                else:
                    femelles_data[femelle]['plaques_9C'].add(plaque_name)
    
    for femelle in sorted(femelles_data.keys()):
        data_fem = femelles_data[femelle]
        nb_croi = len(data_fem['croisements'])
        nb_pl_5c = len(data_fem['plaques_5C'])
        nb_pl_9c = len(data_fem['plaques_9C'])
        plaques_5c_list = ', '.join(sorted(data_fem['plaques_5C'], key=lambda x: int(x.split('_')[1]))[:5])
        plaques_9c_list = ', '.join(sorted(data_fem['plaques_9C'], key=lambda x: int(x.split('_')[1]))[:5])
        
        if nb_pl_5c > 5:
            plaques_5c_list += f" ... (+{nb_pl_5c - 5})"
        if nb_pl_9c > 5:
            plaques_9c_list += f" ... (+{nb_pl_9c - 5})"
        
        writer.writerow([femelle, nb_croi, nb_pl_5c, nb_pl_9c, plaques_5c_list, plaques_9c_list])

print(f"  ✅ {rapport_femelles_file}")

# --- HEATMAP ---
print("\n📊 Génération de la heatmap...")
matrix = pd.DataFrame(0, index=sorted(all_croisements), columns=range(1, nb_plaques_total + 1))

for plaque_name, data in plaques_data.items():
    plaque_num = int(plaque_name.split("_")[1])
    for croi in data['wells'].values():
        if croi in matrix.index:
            matrix.loc[croi, plaque_num] += 1

fig, ax = plt.subplots(figsize=(26, 16))
sns.heatmap(matrix, cmap="YlGnBu", cbar=True, linewidths=0, ax=ax,
            cbar_kws={'label': 'Nombre d\'occurrences'})

ax.axvline(x=100, color='red', linewidth=3, linestyle='--', alpha=0.8)
ax.text(50, -15, '5°C', ha='center', fontsize=16, fontweight='bold', color='green')
ax.text(150, -15, '9°C', ha='center', fontsize=16, fontweight='bold', color='blue')

ax.set_title(f"Répartition avec regroupement par femelle - {total_croisements} croisements sur {len(plaques_data)} plaques",
             fontsize=18, pad=25, fontweight='bold')
ax.set_xlabel("Numéro de plaque", fontsize=14, fontweight='bold')
ax.set_ylabel("Croisements", fontsize=14, fontweight='bold')

plt.tight_layout()
heatmap_path = os.path.join(output_dir, "heatmap_repartition_femelles.png")
plt.savefig(heatmap_path, dpi=300, bbox_inches='tight')
plt.close()
print(f"  ✅ {heatmap_path}")

# --- STATISTIQUES FINALES ---
print("\n" + "="*70)
print("📈 STATISTIQUES FINALES")
print("="*70)
print(f"Groupes traités: {len(groupes)}")
print(f"Croisements traités: {total_croisements}")
print(f"Plaques vérifiées: {len(plaques_data)}")
print(f"  → 5°C: {sum(1 for p in plaques_data.values() if p['temp'] == 5)}")
print(f"  → 9°C: {sum(1 for p in plaques_data.values() if p['temp'] == 9)}")
print(f"\nRegroupement par femelle:")
print(f"  → Moyenne: {avg_femelles:.2f} femelles/plaque")
print(f"  → {adjacence_ok} cas de regroupement correct en colonnes adjacentes")
print(f"\nŒufs par croisement: 16 (8 par température)")
print(f"Total œufs répartis: {total_croisements * 16}")

if not problemes_repartition and not plaques_incompletes:
    print("\n" + "="*70)
    print("✅ VALIDATION RÉUSSIE ! Toutes les vérifications sont OK")
    print("="*70)
else:
    print("\n" + "="*70)
    print("⚠️ ATTENTION : Des problèmes ont été détectés")
    print(f"  • Croisements avec anomalie: {len(problemes_repartition)}")
    print(f"  • Plaques incomplètes: {len(plaques_incompletes)}")
    print("="*70)

print(f"\n📂 Tous les rapports sont dans : {output_dir}")
print("="*70)