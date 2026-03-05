"""
Verifplan_femelles_groupees.py - VÃ‰RIFICATION AVEC REGROUPEMENT PAR FEMELLE

But :
- VÃ©rifier la rÃ©partition correcte des 300 croisements
- VÃ©rifier le regroupement par femelle (croisements BÃ—B + LÃ—B adjacents)
- Chaque croisement : 16 Å“ufs â†’ 4 plaques (8 Ã  5Â°C + 8 Ã  9Â°C)
- 200 plaques totales
- Analyse du regroupement des femelles par plaque
- Lecture directe depuis les fichiers Excel (pas de JSON requis)
"""

import os
import csv
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from collections import defaultdict
import openpyxl

# ---------------- CONFIG ----------------
# CSV source attendu (fichier matrice des croisements, extension .csv)
input_csv = r"A_REMPLACER_PAR_CHEMIN_FICHIER"
# Dossier contenant les fichiers plaques à vérifier (Plaque_001.xlsx ... Plaque_200.xlsx)
plaques_dir = r"A_REMPLACER_PAR_CHEMIN_DOSSIER"
# Dossier de sortie pour les rapports CSV/PNG de vérification
output_dir = plaques_dir  # Les rapports seront dans le mÃªme dossier

if not os.path.exists(plaques_dir):
    print(f"âŒ ERREUR : Dossier non trouvÃ© : {plaques_dir}")
    exit(1)

# Configuration
nb_plaques_total = 200
lignes = ['A', 'B', 'C', 'D']
colonnes = [1, 2, 3, 4, 5, 6]
capacity_per_plaque = 24

print("="*70)
print("ðŸ” VÃ‰RIFICATION DES PLAQUES - REGROUPEMENT PAR FEMELLE")
print("="*70)
print(f"Configuration:")
print(f"  â€¢ Nombre de plaques attendu: {nb_plaques_total}")
print(f"  â€¢ Positions par plaque: {capacity_per_plaque}")
print(f"  â€¢ Dossier: {plaques_dir}")
print("="*70 + "\n")

# --- LECTURE DU CSV SOURCE ---
print("ðŸ“– Lecture du fichier CSV source...")
groupes = []
groupes_noms = []

try:
    with open(input_csv, "r", encoding="utf-8") as f:
        lines = f.readlines()
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            if line and '_G' in line and ',' not in line:
                groupe_name = line
                groupe_croisements = []
                i += 1
                
                if i < len(lines):
                    i += 1
                
                for _ in range(5):
                    if i < len(lines):
                        croi_line = lines[i].strip()
                        if croi_line and ',' in croi_line:
                            parts = croi_line.split(',')
                            for croi in parts[1:6]:
                                croi = croi.strip()
                                if croi and 'x' in croi:
                                    groupe_croisements.append(croi)
                        i += 1
                
                if len(groupe_croisements) == 25:
                    groupes.append(groupe_croisements)
                    groupes_noms.append(groupe_name)
            else:
                i += 1

except FileNotFoundError:
    print(f"âŒ ERREUR : Fichier non trouvÃ© : {input_csv}")
    exit(1)

total_croisements = sum(len(g) for g in groupes)
print(f"âœ… {len(groupes)} groupes extraits")
print(f"âœ… {total_croisements} croisements au total")

all_croisements = [croi for groupe in groupes for croi in groupe]

# --- LECTURE DES PLAQUES EXCEL ---
print("\nðŸ“‚ Lecture des plaques Excel...")

plaques_data = {}
excel_files = [f for f in os.listdir(plaques_dir) if f.startswith("Plaque_") and f.endswith(".xlsx")]

if not excel_files:
    print(f"âŒ ERREUR : Aucun fichier Excel trouvÃ© dans {plaques_dir}")
    exit(1)

print(f"  ðŸ“Š {len(excel_files)} fichiers Excel trouvÃ©s")
print(f"  â³ Lecture en cours...")

for idx, excel_file in enumerate(sorted(excel_files), 1):
    if idx % 20 == 0:
        print(f"     ... {idx}/{len(excel_files)} plaques lues")
    
    plaque_name = excel_file.replace(".xlsx", "")
    file_path = os.path.join(plaques_dir, excel_file)
    
    try:
        # Lire la feuille "Infos" pour la tempÃ©rature
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Lire la tempÃ©rature depuis la feuille Infos
        info_sheet = wb["Infos"]
        temp = None
        for row in info_sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row[0] and "Temperature" in str(row[0]):
                temp_str = str(row[1])
                if "5" in temp_str:
                    temp = 5
                elif "9" in temp_str:
                    temp = 9
                break
        
        # Si pas trouvÃ© dans Infos, dÃ©duire du numÃ©ro de plaque
        if temp is None:
            plaque_num = int(plaque_name.split("_")[1])
            temp = 5 if plaque_num <= 100 else 9
        
        # Lire les croisements depuis la feuille Disposition
        disp_sheet = wb["Disposition"]
        wells = {}
        
        for row_idx, row in enumerate(disp_sheet.iter_rows(min_row=2, max_row=5, values_only=True), 0):
            ligne = lignes[row_idx]
            for col_idx, cell_value in enumerate(row[1:7], 0):  # Colonnes 1-6
                col = colonnes[col_idx]
                pos = f"{ligne}{col}"
                if cell_value and str(cell_value).strip() and 'x' in str(cell_value):
                    wells[pos] = str(cell_value).strip()
        
        plaques_data[plaque_name] = {
            'temp': temp,
            'wells': wells
        }
        
        wb.close()
        
    except Exception as e:
        print(f"  âš ï¸ Erreur lors de la lecture de {excel_file}: {e}")
        continue

print(f"  âœ… {len(plaques_data)} plaques chargÃ©es avec succÃ¨s\n")

if len(plaques_data) == 0:
    print("âŒ ERREUR : Aucune plaque n'a pu Ãªtre lue")
    exit(1)

# --- VÃ‰RIFICATIONS GLOBALES ---
print("ðŸ” VÃ©rifications globales...")

# 1. Nombre de plaques
if len(plaques_data) != nb_plaques_total:
    print(f"  âš ï¸ {len(plaques_data)} plaques trouvÃ©es (attendu: {nb_plaques_total})")
else:
    print(f"  âœ… {nb_plaques_total} plaques prÃ©sentes")

# 2. Compter les occurrences par croisement
occ_by_crois = {c: {"5Â°C": 0, "9Â°C": 0, "Total": 0} for c in all_croisements}

for plaque_name, data in plaques_data.items():
    temp_key = "5Â°C" if data['temp'] == 5 else "9Â°C"
    for croi in data['wells'].values():
        if croi in occ_by_crois:
            occ_by_crois[croi][temp_key] += 1
            occ_by_crois[croi]["Total"] += 1
        else:
            # Croisement non attendu
            if croi not in occ_by_crois:
                occ_by_crois[croi] = {"5Â°C": 0, "9Â°C": 0, "Total": 0}
            occ_by_crois[croi][temp_key] += 1
            occ_by_crois[croi]["Total"] += 1

# 3. VÃ©rifier les anomalies de rÃ©partition
problemes_repartition = []
for croi, counts in occ_by_crois.items():
    if counts["5Â°C"] != 8 or counts["9Â°C"] != 8 or counts["Total"] != 16:
        problemes_repartition.append({
            'croisement': croi,
            '5C': counts["5Â°C"],
            '9C': counts["9Â°C"],
            'total': counts["Total"]
        })

if problemes_repartition:
    print(f"  âŒ {len(problemes_repartition)} croisements avec anomalie de rÃ©partition")
    for p in problemes_repartition[:5]:  # Afficher les 5 premiers
        print(f"     â€¢ {p['croisement']}: {p['5C']}+{p['9C']}={p['total']} (attendu: 8+8=16)")
    if len(problemes_repartition) > 5:
        print(f"     ... et {len(problemes_repartition) - 5} autres")
else:
    print(f"  âœ… Tous les croisements correctement rÃ©partis (8+8=16 Å“ufs)")

# 4. VÃ©rifier le remplissage des plaques
plaques_incompletes = []
for plaque_name, data in plaques_data.items():
    if len(data['wells']) != capacity_per_plaque:
        plaques_incompletes.append({
            'plaque': plaque_name,
            'positions': len(data['wells'])
        })

if plaques_incompletes:
    print(f"  âš ï¸ {len(plaques_incompletes)} plaques incomplÃ¨tes")
    for p in plaques_incompletes[:5]:
        print(f"     â€¢ {p['plaque']}: {p['positions']}/24 positions")
    if len(plaques_incompletes) > 5:
        print(f"     ... et {len(plaques_incompletes) - 5} autres")
else:
    print(f"  âœ… Toutes les plaques complÃ¨tes (24/24 positions)")

# --- ANALYSE DU REGROUPEMENT PAR FEMELLE ---
print("\nðŸ‘©â€ðŸ”¬ Analyse du regroupement par femelle...")

# Extraire les femelles de chaque croisement
def extract_female(croisement):
    """Extrait la femelle d'un croisement (aprÃ¨s le 'x')"""
    if 'x' in croisement:
        return croisement.split('x')[1]
    return None

# Analyser chaque plaque
plaques_regroupement = {}
for plaque_name, data in plaques_data.items():
    # Compter les croisements par femelle dans cette plaque
    femelles_count = defaultdict(lambda: {'positions': [], 'croisements': []})
    
    for pos, croi in data['wells'].items():
        femelle = extract_female(croi)
        if femelle:
            femelles_count[femelle]['positions'].append(pos)
            femelles_count[femelle]['croisements'].append(croi)
    
    plaques_regroupement[plaque_name] = {
        'temp': data['temp'],
        'femelles': dict(femelles_count),
        'nb_femelles': len(femelles_count)
    }

# Statistiques sur le regroupement
femelles_par_plaque = [info['nb_femelles'] for info in plaques_regroupement.values()]
avg_femelles = sum(femelles_par_plaque) / len(femelles_par_plaque) if femelles_par_plaque else 0

print(f"  ðŸ“Š Nombre moyen de femelles par plaque: {avg_femelles:.2f}")
print(f"  ðŸ“Š Min: {min(femelles_par_plaque)} femelles, Max: {max(femelles_par_plaque)} femelles")

# VÃ©rifier l'adjacence des colonnes pour une mÃªme femelle
adjacence_ok = 0
adjacence_problemes = []

for plaque_name, info in plaques_regroupement.items():
    for femelle, data_fem in info['femelles'].items():
        positions = data_fem['positions']
        # Extraire les colonnes
        cols = sorted(set([int(pos[1]) for pos in positions]))
        
        # VÃ©rifier si les colonnes sont adjacentes
        if len(cols) > 1:
            is_adjacent = all(cols[i+1] - cols[i] == 1 for i in range(len(cols)-1))
            if is_adjacent:
                adjacence_ok += 1
            else:
                adjacence_problemes.append({
                    'plaque': plaque_name,
                    'femelle': femelle,
                    'colonnes': cols,
                    'croisements': data_fem['croisements']
                })

if adjacence_problemes:
    print(f"  âš ï¸ {len(adjacence_problemes)} cas oÃ¹ les colonnes d'une femelle ne sont pas adjacentes")
    for p in adjacence_problemes[:3]:
        print(f"     â€¢ {p['plaque']} - {p['femelle']}: colonnes {p['colonnes']}")
    if len(adjacence_problemes) > 3:
        print(f"     ... et {len(adjacence_problemes) - 3} autres")
else:
    print(f"  âœ… Toutes les femelles avec plusieurs colonnes sont bien regroupÃ©es en colonnes adjacentes")

# --- ANALYSE PAR TYPE DE CROISEMENT ---
print("\nðŸ§¬ Analyse par type de croisement...")

types_croisements = {
    'BÃ—B': [],
    'LÃ—L': [],
    'LÃ—B': [],
    'BÃ—L': []
}

for croi in all_croisements:
    if 'B_M' in croi and 'B_F' in croi:
        types_croisements['BÃ—B'].append(croi)
    elif 'L_M' in croi and 'L_F' in croi:
        types_croisements['LÃ—L'].append(croi)
    elif 'L_M' in croi and 'B_F' in croi:
        types_croisements['LÃ—B'].append(croi)
    elif 'B_M' in croi and 'L_F' in croi:
        types_croisements['BÃ—L'].append(croi)

for type_nom, croisements_type in types_croisements.items():
    print(f"  â€¢ {type_nom}: {len(croisements_type)} croisements")

# --- RAPPORTS CSV ---
print("\nðŸ“‹ GÃ©nÃ©ration des rapports...")

# 1. Rapport croisements
rapport_crois_file = os.path.join(output_dir, "rapport_croisements_femelles.csv")
with open(rapport_crois_file, "w", newline='', encoding='utf-8') as csvf:
    writer = csv.writer(csvf)
    writer.writerow(["Croisement", "Femelle", "Type", "Occ_5Â°C", "Occ_9Â°C", "Total", "Attendu", "Statut"])
    
    for c in sorted(occ_by_crois.keys()):
        femelle = extract_female(c)
        # DÃ©terminer le type
        if 'B_M' in c and 'B_F' in c:
            type_croi = "BÃ—B"
        elif 'L_M' in c and 'L_F' in c:
            type_croi = "LÃ—L"
        elif 'L_M' in c and 'B_F' in c:
            type_croi = "LÃ—B"
        elif 'B_M' in c and 'L_F' in c:
            type_croi = "BÃ—L"
        else:
            type_croi = "Inconnu"
        
        a = occ_by_crois[c]["5Â°C"]
        b = occ_by_crois[c]["9Â°C"]
        total = occ_by_crois[c]["Total"]
        status = "OK" if (a == 8 and b == 8 and total == 16) else "ERREUR"
        writer.writerow([c, femelle, type_croi, a, b, total, 16, status])

print(f"  âœ… {rapport_crois_file}")

# 2. Rapport plaques avec analyse femelles
rapport_plaques_file = os.path.join(output_dir, "rapport_plaques_femelles.csv")
with open(rapport_plaques_file, "w", newline='', encoding='utf-8') as csvf:
    writer = csv.writer(csvf)
    writer.writerow(["Plaque", "TempÃ©rature", "Nb_positions", "Nb_femelles", "Femelles_prÃ©sentes", "Statut"])
    
    for plaque_name in sorted(plaques_data.keys(), key=lambda x: int(x.split("_")[1])):
        data = plaques_data[plaque_name]
        info_regr = plaques_regroupement[plaque_name]
        
        temp = f"{data['temp']}Â°C"
        nb_pos = len(data['wells'])
        nb_fem = info_regr['nb_femelles']
        femelles_list = ', '.join(sorted(info_regr['femelles'].keys()))
        
        status = "OK" if nb_pos == capacity_per_plaque else "ERREUR"
        writer.writerow([plaque_name, temp, nb_pos, nb_fem, femelles_list, status])

print(f"  âœ… {rapport_plaques_file}")

# 3. Rapport dÃ©taillÃ© par femelle
rapport_femelles_file = os.path.join(output_dir, "rapport_par_femelle.csv")
with open(rapport_femelles_file, "w", newline='', encoding='utf-8') as csvf:
    writer = csv.writer(csvf)
    writer.writerow(["Femelle", "Nb_croisements_total", "Nb_plaques_5C", "Nb_plaques_9C", "Plaques_5C", "Plaques_9C"])
    
    # Regrouper par femelle
    femelles_data = defaultdict(lambda: {'croisements': set(), 'plaques_5C': set(), 'plaques_9C': set()})
    
    for plaque_name, data in plaques_data.items():
        for croi in data['wells'].values():
            femelle = extract_female(croi)
            if femelle:
                femelles_data[femelle]['croisements'].add(croi)
                if data['temp'] == 5:
                    femelles_data[femelle]['plaques_5C'].add(plaque_name)
                else:
                    femelles_data[femelle]['plaques_9C'].add(plaque_name)
    
    for femelle in sorted(femelles_data.keys()):
        data_fem = femelles_data[femelle]
        nb_croi = len(data_fem['croisements'])
        nb_pl_5c = len(data_fem['plaques_5C'])
        nb_pl_9c = len(data_fem['plaques_9C'])
        plaques_5c_list = ', '.join(sorted(data_fem['plaques_5C'], key=lambda x: int(x.split('_')[1]))[:5])
        plaques_9c_list = ', '.join(sorted(data_fem['plaques_9C'], key=lambda x: int(x.split('_')[1]))[:5])
        
        if nb_pl_5c > 5:
            plaques_5c_list += f" ... (+{nb_pl_5c - 5})"
        if nb_pl_9c > 5:
            plaques_9c_list += f" ... (+{nb_pl_9c - 5})"
        
        writer.writerow([femelle, nb_croi, nb_pl_5c, nb_pl_9c, plaques_5c_list, plaques_9c_list])

print(f"  âœ… {rapport_femelles_file}")

# --- HEATMAP ---
print("\nðŸ“Š GÃ©nÃ©ration de la heatmap...")
matrix = pd.DataFrame(0, index=sorted(all_croisements), columns=range(1, nb_plaques_total + 1))

for plaque_name, data in plaques_data.items():
    plaque_num = int(plaque_name.split("_")[1])
    for croi in data['wells'].values():
        if croi in matrix.index:
            matrix.loc[croi, plaque_num] += 1

fig, ax = plt.subplots(figsize=(26, 16))
sns.heatmap(matrix, cmap="YlGnBu", cbar=True, linewidths=0, ax=ax,
            cbar_kws={'label': 'Nombre d\'occurrences'})

ax.axvline(x=100, color='red', linewidth=3, linestyle='--', alpha=0.8)
ax.text(50, -15, '5Â°C', ha='center', fontsize=16, fontweight='bold', color='green')
ax.text(150, -15, '9Â°C', ha='center', fontsize=16, fontweight='bold', color='blue')

ax.set_title(f"RÃ©partition avec regroupement par femelle - {total_croisements} croisements sur {len(plaques_data)} plaques",
             fontsize=18, pad=25, fontweight='bold')
ax.set_xlabel("NumÃ©ro de plaque", fontsize=14, fontweight='bold')
ax.set_ylabel("Croisements", fontsize=14, fontweight='bold')

plt.tight_layout()
heatmap_path = os.path.join(output_dir, "heatmap_repartition_femelles.png")
plt.savefig(heatmap_path, dpi=300, bbox_inches='tight')
plt.close()
print(f"  âœ… {heatmap_path}")

# --- STATISTIQUES FINALES ---
print("\n" + "="*70)
print("ðŸ“ˆ STATISTIQUES FINALES")
print("="*70)
print(f"Groupes traitÃ©s: {len(groupes)}")
print(f"Croisements traitÃ©s: {total_croisements}")
print(f"Plaques vÃ©rifiÃ©es: {len(plaques_data)}")
print(f"  â†’ 5Â°C: {sum(1 for p in plaques_data.values() if p['temp'] == 5)}")
print(f"  â†’ 9Â°C: {sum(1 for p in plaques_data.values() if p['temp'] == 9)}")
print(f"\nRegroupement par femelle:")
print(f"  â†’ Moyenne: {avg_femelles:.2f} femelles/plaque")
print(f"  â†’ {adjacence_ok} cas de regroupement correct en colonnes adjacentes")
print(f"\nÅ’ufs par croisement: 16 (8 par tempÃ©rature)")
print(f"Total Å“ufs rÃ©partis: {total_croisements * 16}")

if not problemes_repartition and not plaques_incompletes:
    print("\n" + "="*70)
    print("âœ… VALIDATION RÃ‰USSIE ! Toutes les vÃ©rifications sont OK")
    print("="*70)
else:
    print("\n" + "="*70)
    print("âš ï¸ ATTENTION : Des problÃ¨mes ont Ã©tÃ© dÃ©tectÃ©s")
    print(f"  â€¢ Croisements avec anomalie: {len(problemes_repartition)}")
    print(f"  â€¢ Plaques incomplÃ¨tes: {len(plaques_incompletes)}")
    print("="*70)

print(f"\nðŸ“‚ Tous les rapports sont dans : {output_dir}")
print("="*70)
