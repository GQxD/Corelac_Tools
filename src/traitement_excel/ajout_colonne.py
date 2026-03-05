import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill
from copy import copy

def ajouter_colonnes_excel(dossier_source, dossier_destination=None, nb_dates_photo=5):
    """
    Ajoute les colonnes 'Tray ID' et plusieurs colonnes 'Date photo X' 
    Ã  la feuille 'Suivi' de tous les fichiers Excel dans un dossier.
    
    Args:
        dossier_source: Chemin du dossier contenant les fichiers Excel
        dossier_destination: Chemin du dossier pour les fichiers modifiÃ©s 
                            (optionnel, par dÃ©faut crÃ©e un sous-dossier 'modifiÃ©s')
        nb_dates_photo: Nombre de colonnes pour les dates de photographie (dÃ©faut: 5)
    """
    # VÃ©rifier que le dossier source existe
    if not os.path.exists(dossier_source):
        print(f"âŒ ERREUR: Le dossier '{dossier_source}' n'existe pas!")
        return
    
    # Lister tous les fichiers du dossier pour diagnostic
    print(f"ðŸ“ Contenu du dossier:")
    tous_fichiers = os.listdir(dossier_source)
    
    # Filtrer uniquement les fichiers Plaque_001 Ã  Plaque_200
    fichiers_excel = []
    for f in tous_fichiers:
        if re.match(r'Plaque_\d{3}\.(xlsx|xls)$', f, re.IGNORECASE):
            fichiers_excel.append(f)
    
    fichiers_excel.sort()  # Trier par ordre numÃ©rique
    
    print(f"   Total de fichiers: {len(tous_fichiers)}")
    print(f"   Fichiers Plaque_XXX trouvÃ©s: {len(fichiers_excel)}")
    
    if len(fichiers_excel) == 0:
        print("\nâš ï¸  Aucun fichier Plaque_XXX.xlsx trouvÃ©!")
        print("   Le script cherche les fichiers nommÃ©s: Plaque_001.xlsx, Plaque_002.xlsx, etc.")
        print("   VÃ©rifiez que:")
        print("   - Les fichiers sont bien nommÃ©s avec le format Plaque_XXX.xlsx")
        print("   - Les numÃ©ros ont 3 chiffres (001, 002, etc.)")
        print("\n   Exemples de fichiers dans le dossier:")
        for f in tous_fichiers[:10]:
            print(f"   - {f}")
        return
    
    print(f"\n   Fichiers Plaque Ã  traiter:")
    for f in fichiers_excel[:5]:
        print(f"   - {f}")
    if len(fichiers_excel) > 5:
        print(f"   ... et {len(fichiers_excel) - 5} autres")
    print(f"\n   Configuration: Tray ID + {nb_dates_photo} colonnes de dates photo")
    print()
    
    # CrÃ©er le dossier de destination s'il n'existe pas
    if dossier_destination is None:
        dossier_destination = os.path.join(dossier_source, "modifiÃ©s")
    
    if not os.path.exists(dossier_destination):
        os.makedirs(dossier_destination)
    
    # Compter les fichiers traitÃ©s
    fichiers_traites = 0
    fichiers_erreur = 0
    
    # Parcourir tous les fichiers Excel
    for fichier in fichiers_excel:
        chemin_fichier = os.path.join(dossier_source, fichier)
        
        try:
            print(f"Traitement de: {fichier}...")
            
            # Charger le fichier Excel
            wb = load_workbook(chemin_fichier)
            
            # Afficher les feuilles disponibles
            print(f"   Feuilles disponibles: {', '.join(wb.sheetnames)}")
            
            # VÃ©rifier si la feuille 'Suivi' existe
            if 'Suivi' not in wb.sheetnames:
                print(f"  âš ï¸  Feuille 'Suivi' non trouvÃ©e dans {fichier}")
                print(f"      Utilisez une de ces feuilles: {', '.join(wb.sheetnames)}")
                fichiers_erreur += 1
                continue
            
            # SÃ©lectionner la feuille 'Suivi'
            ws = wb['Suivi']
            
            # Trouver la derniÃ¨re colonne utilisÃ©e
            derniere_colonne = ws.max_column
            
            # Copier le style de la derniÃ¨re colonne existante (en-tÃªte)
            cellule_reference = ws.cell(row=1, column=derniere_colonne)
            
            # Ajouter la colonne 'Tray ID'
            nouvelle_col = derniere_colonne + 1
            cellule_tray = ws.cell(row=1, column=nouvelle_col, value='Tray ID')
            
            # Copier le style
            if cellule_reference.font:
                cellule_tray.font = copy(cellule_reference.font)
            if cellule_reference.border:
                cellule_tray.border = copy(cellule_reference.border)
            if cellule_reference.fill:
                cellule_tray.fill = copy(cellule_reference.fill)
            if cellule_reference.alignment:
                cellule_tray.alignment = copy(cellule_reference.alignment)
            
            # Ajouter les colonnes de dates photo
            for i in range(1, nb_dates_photo + 1):
                col_index = nouvelle_col + i
                cellule_date = ws.cell(row=1, column=col_index, value=f'Date photo {i}')
                
                # Copier le style
                if cellule_reference.font:
                    cellule_date.font = copy(cellule_reference.font)
                if cellule_reference.border:
                    cellule_date.border = copy(cellule_reference.border)
                if cellule_reference.fill:
                    cellule_date.fill = copy(cellule_reference.fill)
                if cellule_reference.alignment:
                    cellule_date.alignment = copy(cellule_reference.alignment)
            
            # Ajuster la largeur des nouvelles colonnes (optionnel)
            for col in range(nouvelle_col, nouvelle_col + nb_dates_photo + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
            # Sauvegarder le fichier modifiÃ©
            chemin_destination = os.path.join(dossier_destination, fichier)
            wb.save(chemin_destination)
            
            print(f"  âœ“ {fichier} traitÃ© avec succÃ¨s")
            fichiers_traites += 1
            
        except Exception as e:
            print(f"  âœ— Erreur avec {fichier}: {str(e)}")
            fichiers_erreur += 1
    
    # RÃ©sumÃ©
    print("\n" + "="*50)
    print(f"Traitement terminÃ© !")
    print(f"âœ“ Fichiers traitÃ©s: {fichiers_traites}")
    print(f"âœ— Fichiers en erreur: {fichiers_erreur}")
    print(f"ðŸ“ Fichiers sauvegardÃ©s dans: {dossier_destination}")
    print("="*50)


if __name__ == "__main__":
    # Dossier source attendu: contient des fichiers Excel de plaques (Plaque_XXX.xlsx)
    dossier_source = r"A_REMPLACER_PAR_CHEMIN_DOSSIER"
    
    # MODIFIEZ LE NOMBRE DE COLONNES POUR LES DATES PHOTO (dÃ©faut: 5)
    nb_dates_photo = 50  # Changez ce nombre selon vos besoins
    
    # Optionnel: dossier de sortie (sinon un sous-dossier local sera créé automatiquement)
    # dossier_destination = r"A_REMPLACER_PAR_CHEMIN_DOSSIER"
    
    print("DÃ©but du traitement des fichiers Excel...")
    print(f"Dossier source: {dossier_source}\n")
    
    ajouter_colonnes_excel(dossier_source, nb_dates_photo=nb_dates_photo)
    
    print("\nAppuyez sur EntrÃ©e pour fermer...")
    input()
