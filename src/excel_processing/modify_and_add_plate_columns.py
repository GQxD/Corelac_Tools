import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill
from copy import copy

def modifier_et_ajouter_colonnes_plaques(dossier_source, dossier_destination=None, nb_dates_photo=50):
    """
    Modifie les croisements ET ajoute les colonnes 'Tray ID' et 'Date photo X' 
    dans les fichiers Plaque_001 Ã  Plaque_200.
    
    Modifications appliquÃ©es:
    1. Remplacement des femelles B_F11 Ã  B_F15 selon nouvelle matrice (en gras)
    2. Interversion B_M1 â†” B_M7 pour croisements BOURGETÃ—BOURGET (en italique)
    3. Ajout colonne 'Tray ID' dans feuille 'Suivi'
    4. Ajout de X colonnes 'Date photo X' dans feuille 'Suivi'
    
    Args:
        dossier_source: Chemin du dossier contenant les fichiers Excel
        dossier_destination: Chemin du dossier pour les fichiers modifiÃ©s
        nb_dates_photo: Nombre de colonnes pour les dates de photographie (dÃ©faut: 50)
    """
    
    # Dictionnaire de remplacement pour les femelles B_F11 Ã  B_F15
    remplacements_femelles = {
        # Bourget Ã— Bourget - MÃ¢les M11, M12, M13
        'B_M11xB_F11': 'B_M11xB_F1', 'B_M11xB_F12': 'B_M11xB_F2', 'B_M11xB_F13': 'B_M11xB_F3',
        'B_M11xB_F14': 'B_M11xB_F4', 'B_M11xB_F15': 'B_M11xB_F5',
        
        'B_M12xB_F11': 'B_M12xB_F1', 'B_M12xB_F12': 'B_M12xB_F2', 'B_M12xB_F13': 'B_M12xB_F3',
        'B_M12xB_F14': 'B_M12xB_F4', 'B_M12xB_F15': 'B_M12xB_F5',
        
        'B_M13xB_F11': 'B_M13xB_F1', 'B_M13xB_F12': 'B_M13xB_F2', 'B_M13xB_F13': 'B_M13xB_F3',
        'B_M13xB_F14': 'B_M13xB_F4', 'B_M13xB_F15': 'B_M13xB_F5',
        
        # Bourget Ã— Bourget - MÃ¢les M14, M15
        'B_M14xB_F11': 'B_M14xB_F6', 'B_M14xB_F12': 'B_M14xB_F7', 'B_M14xB_F13': 'B_M14xB_F8',
        'B_M14xB_F14': 'B_M14xB_F9', 'B_M14xB_F15': 'B_M14xB_F10',
        
        'B_M15xB_F11': 'B_M15xB_F6', 'B_M15xB_F12': 'B_M15xB_F7', 'B_M15xB_F13': 'B_M15xB_F8',
        'B_M15xB_F14': 'B_M15xB_F9', 'B_M15xB_F15': 'B_M15xB_F10',
        
        # LÃ©man Ã— Bourget - MÃ¢les M11, M12, M13
        'L_M11xB_F11': 'L_M11xB_F1', 'L_M11xB_F12': 'L_M11xB_F2', 'L_M11xB_F13': 'L_M11xB_F3',
        'L_M11xB_F14': 'L_M11xB_F4', 'L_M11xB_F15': 'L_M11xB_F5',
        
        'L_M12xB_F11': 'L_M12xB_F1', 'L_M12xB_F12': 'L_M12xB_F2', 'L_M12xB_F13': 'L_M12xB_F3',
        'L_M12xB_F14': 'L_M12xB_F4', 'L_M12xB_F15': 'L_M12xB_F5',
        
        'L_M13xB_F11': 'L_M13xB_F1', 'L_M13xB_F12': 'L_M13xB_F2', 'L_M13xB_F13': 'L_M13xB_F3',
        'L_M13xB_F14': 'L_M13xB_F4', 'L_M13xB_F15': 'L_M13xB_F5',
        
        # LÃ©man Ã— Bourget - MÃ¢les M14, M15
        'L_M14xB_F11': 'L_M14xB_F6', 'L_M14xB_F12': 'L_M14xB_F7', 'L_M14xB_F13': 'L_M14xB_F8',
        'L_M14xB_F14': 'L_M14xB_F9', 'L_M14xB_F15': 'L_M14xB_F10',
        
        'L_M15xB_F11': 'L_M15xB_F6', 'L_M15xB_F12': 'L_M15xB_F7', 'L_M15xB_F13': 'L_M15xB_F8',
        'L_M15xB_F14': 'L_M15xB_F9', 'L_M15xB_F15': 'L_M15xB_F10',
    }
    
    # VÃ©rifier que le dossier source existe
    if not os.path.exists(dossier_source):
        print(f"âŒ ERREUR: Le dossier '{dossier_source}' n'existe pas!")
        return
    
    # Lister tous les fichiers du dossier
    print(f"ðŸ“ Contenu du dossier:")
    tous_fichiers = os.listdir(dossier_source)
    
    # Filtrer uniquement les fichiers Plaque_XXX
    fichiers_excel = []
    for f in tous_fichiers:
        if re.match(r'Plaque_\d{3}\.(xlsx|xls)$', f, re.IGNORECASE):
            fichiers_excel.append(f)
    
    fichiers_excel.sort()
    
    print(f"   Total de fichiers: {len(tous_fichiers)}")
    print(f"   Fichiers Plaque_XXX trouvÃ©s: {len(fichiers_excel)}")
    
    if len(fichiers_excel) == 0:
        print("\nâš ï¸  Aucun fichier Plaque_XXX.xlsx trouvÃ©!")
        print("   Le script cherche les fichiers nommÃ©s: Plaque_001.xlsx, Plaque_002.xlsx, etc.")
        print("   VÃ©rifiez que:")
        print("   - Les fichiers sont bien nommÃ©s avec le format Plaque_XXX.xlsx")
        print("   - Les numÃ©ros ont 3 chiffres (001, 002, etc.)")
        return
    
    print(f"\n   Fichiers Ã  traiter:")
    for f in fichiers_excel[:5]:
        print(f"   - {f}")
    if len(fichiers_excel) > 5:
        print(f"   ... et {len(fichiers_excel) - 5} autres")
    
    print(f"\n   Configuration: Tray ID + {nb_dates_photo} colonnes de dates photo")
    print()
    
    # CrÃ©er le dossier de destination
    if dossier_destination is None:
        dossier_destination = os.path.join(dossier_source, "plaques_complÃ¨tes")
    
    if not os.path.exists(dossier_destination):
        os.makedirs(dossier_destination)
    
    print("ðŸ”§ Modifications Ã  appliquer:")
    print("  1. Remplacement femelles B_F11-F15 â†’ nouvelles femelles (GRAS)")
    print("  2. Interversion B_M1 â†” B_M7 pour BOURGETÃ—BOURGET (ITALIQUE)")
    print(f"  3. Ajout colonne 'Tray ID' dans feuille 'Suivi'")
    print(f"  4. Ajout de {nb_dates_photo} colonnes 'Date photo X' dans feuille 'Suivi'")
    print()
    
    fichiers_traites = 0
    fichiers_erreur = 0
    total_modifs_femelles = 0
    total_interversions = 0
    fichiers_avec_suivi = 0
    
    for fichier in fichiers_excel:
        chemin_fichier = os.path.join(dossier_source, fichier)
        nom_plaque = os.path.splitext(fichier)[0]
        
        try:
            print(f"Traitement: {fichier}...")
            
            # Charger le fichier Excel
            wb = load_workbook(chemin_fichier)
            
            modifs_femelles_plaque = 0
            interversions_plaque = 0
            
            # ===== PARTIE 1 & 2: MODIFICATION DES CROISEMENTS =====
            # Traiter toutes les feuilles pour les modifications de croisements
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            valeur_originale = cell.value
                            nouvelle_valeur = valeur_originale
                            appliquer_gras = False
                            appliquer_italique = False
                            
                            # Modification 1: Remplacement des femelles B_F11-F15
                            if valeur_originale in remplacements_femelles:
                                nouvelle_valeur = remplacements_femelles[valeur_originale]
                                appliquer_gras = True
                                modifs_femelles_plaque += 1
                            
                            # Modification 2: Interversion B_M1 â†” B_M7 (uniquement BOURGETÃ—BOURGET)
                            if nouvelle_valeur.startswith('B_M1x') and 'B_F' in nouvelle_valeur:
                                nouvelle_valeur = nouvelle_valeur.replace('B_M1x', 'B_M7x')
                                appliquer_italique = True
                                interversions_plaque += 1
                            elif nouvelle_valeur.startswith('B_M7x') and 'B_F' in nouvelle_valeur:
                                nouvelle_valeur = nouvelle_valeur.replace('B_M7x', 'B_M1x')
                                appliquer_italique = True
                                interversions_plaque += 1
                            
                            # Appliquer les modifications
                            if nouvelle_valeur != valeur_originale or appliquer_gras or appliquer_italique:
                                cell.value = nouvelle_valeur
                                
                                # Copier le style existant ou crÃ©er un nouveau
                                if cell.font:
                                    nouvelle_font = copy(cell.font)
                                else:
                                    nouvelle_font = Font()
                                
                                # Appliquer le formatage
                                if appliquer_gras:
                                    nouvelle_font = Font(
                                        name=nouvelle_font.name,
                                        size=nouvelle_font.size,
                                        bold=True,
                                        italic=nouvelle_font.italic or appliquer_italique,
                                        color=nouvelle_font.color
                                    )
                                if appliquer_italique:
                                    nouvelle_font = Font(
                                        name=nouvelle_font.name,
                                        size=nouvelle_font.size,
                                        bold=nouvelle_font.bold or appliquer_gras,
                                        italic=True,
                                        color=nouvelle_font.color
                                    )
                                
                                cell.font = nouvelle_font
            
            # ===== PARTIE 3 & 4: AJOUT DES COLONNES DANS FEUILLE 'SUIVI' =====
            if 'Suivi' in wb.sheetnames:
                ws_suivi = wb['Suivi']
                
                # Trouver la derniÃ¨re colonne utilisÃ©e
                derniere_colonne = ws_suivi.max_column
                
                # Copier le style de la derniÃ¨re colonne existante (en-tÃªte)
                cellule_reference = ws_suivi.cell(row=1, column=derniere_colonne)
                
                # Ajouter la colonne 'Tray ID'
                nouvelle_col = derniere_colonne + 1
                cellule_tray = ws_suivi.cell(row=1, column=nouvelle_col, value='Tray ID')
                
                # Copier le style
                if cellule_reference.font:
                    cellule_tray.font = copy(cellule_reference.font)
                if cellule_reference.border:
                    cellule_tray.border = copy(cellule_reference.border)
                if cellule_reference.fill:
                    cellule_tray.fill = copy(cellule_reference.fill)
                if cellule_reference.alignment:
                    cellule_tray.alignment = copy(cellule_reference.alignment)
                
                # Ajouter les colonnes de dates photo
                for i in range(1, nb_dates_photo + 1):
                    col_index = nouvelle_col + i
                    cellule_date = ws_suivi.cell(row=1, column=col_index, value=f'Date photo {i}')
                    
                    # Copier le style
                    if cellule_reference.font:
                        cellule_date.font = copy(cellule_reference.font)
                    if cellule_reference.border:
                        cellule_date.border = copy(cellule_reference.border)
                    if cellule_reference.fill:
                        cellule_date.fill = copy(cellule_reference.fill)
                    if cellule_reference.alignment:
                        cellule_date.alignment = copy(cellule_reference.alignment)
                
                # Ajuster la largeur des nouvelles colonnes
                for col in range(nouvelle_col, nouvelle_col + nb_dates_photo + 1):
                    ws_suivi.column_dimensions[ws_suivi.cell(row=1, column=col).column_letter].width = 15
                
                fichiers_avec_suivi += 1
                print(f"  âœ“ Feuille 'Suivi': Tray ID + {nb_dates_photo} colonnes ajoutÃ©es")
            else:
                print(f"  âš ï¸  Feuille 'Suivi' non trouvÃ©e (colonnes non ajoutÃ©es)")
            
            # Sauvegarder le fichier modifiÃ©
            chemin_destination = os.path.join(dossier_destination, fichier)
            wb.save(chemin_destination)
            
            print(f"  âœ“ ModifiÃ©: {modifs_femelles_plaque} remplacements femelles, {interversions_plaque} interversions")
            fichiers_traites += 1
            total_modifs_femelles += modifs_femelles_plaque
            total_interversions += interversions_plaque
            
        except Exception as e:
            print(f"  âœ— Erreur avec {fichier}: {str(e)}")
            fichiers_erreur += 1
    
    # RÃ©sumÃ©
    print("\n" + "="*70)
    print(f"Traitement terminÃ© !")
    print(f"âœ“ Fichiers traitÃ©s avec succÃ¨s: {fichiers_traites}/{len(fichiers_excel)}")
    print(f"âœ— Fichiers en erreur: {fichiers_erreur}")
    print()
    print(f"ðŸ“Š Modifications des croisements:")
    print(f"   - Total remplacements femelles (GRAS): {total_modifs_femelles}")
    print(f"   - Total interversions B_M1â†”B_M7 (ITALIQUE): {total_interversions}")
    print()
    print(f"ðŸ“‹ Ajout de colonnes:")
    print(f"   - Fichiers avec feuille 'Suivi' modifiÃ©e: {fichiers_avec_suivi}")
    print(f"   - Colonnes ajoutÃ©es par fichier: Tray ID + {nb_dates_photo} dates photo")
    print()
    print(f"ðŸ“ Fichiers sauvegardÃ©s dans: {dossier_destination}")
    print("="*70)


if __name__ == "__main__":
    # Dossier source attendu: contient des fichiers Excel de plaques (Plaque_XXX.xlsx)
    dossier_source = r"A_REMPLACER_PAR_CHEMIN_DOSSIER"
    
    # MODIFIEZ LE NOMBRE DE COLONNES POUR LES DATES PHOTO (dÃ©faut: 50)
    nb_dates_photo = 50
    
    # Optionnel: dossier de sortie (sinon un sous-dossier local sera créé automatiquement)
    # dossier_destination = r"A_REMPLACER_PAR_CHEMIN_DOSSIER"
    
    print("ðŸš€ DÃ©but du traitement complet des plaques...")
    print(f"ðŸ“ Dossier source: {dossier_source}")
    print(f"ðŸ“¸ Nombre de colonnes photos: {nb_dates_photo}\n")
    
    modifier_et_ajouter_colonnes_plaques(dossier_source, nb_dates_photo=nb_dates_photo)
    
    print("\nAppuyez sur EntrÃ©e pour fermer...")
    input()
