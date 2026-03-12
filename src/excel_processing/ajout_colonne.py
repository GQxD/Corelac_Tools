import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill
from copy import copy

def ajouter_colonnes_excel(dossier_source, dossier_destination=None, nb_dates_photo=5):
    """
    Ajoute les colonnes 'Tray ID' et plusieurs colonnes 'Date photo X' 
    à la feuille 'Suivi' de tous les fichiers Excel dans un dossier.
    
    Args:
        dossier_source: Chemin du dossier contenant les fichiers Excel
        dossier_destination: Chemin du dossier pour les fichiers modifiés 
                            (optionnel, par défaut crée un sous-dossier 'modifiés')
        nb_dates_photo: Nombre de colonnes pour les dates de photographie (défaut: 5)
    """
    # Vérifier que le dossier source existe
    if not os.path.exists(dossier_source):
        print(f"❌ ERREUR: Le dossier '{dossier_source}' n'existe pas!")
        return
    
    # Lister tous les fichiers du dossier pour diagnostic
    print(f"📁 Contenu du dossier:")
    tous_fichiers = os.listdir(dossier_source)
    
    # Filtrer uniquement les fichiers Plaque_001 à Plaque_200
    fichiers_excel = []
    for f in tous_fichiers:
        if re.match(r'Plaque_\d{3}\.(xlsx|xls)$', f, re.IGNORECASE):
            fichiers_excel.append(f)
    
    fichiers_excel.sort()  # Trier par ordre numérique
    
    print(f"   Total de fichiers: {len(tous_fichiers)}")
    print(f"   Fichiers Plaque_XXX trouvés: {len(fichiers_excel)}")
    
    if len(fichiers_excel) == 0:
        print("\n⚠️  Aucun fichier Plaque_XXX.xlsx trouvé!")
        print("   Le script cherche les fichiers nommés: Plaque_001.xlsx, Plaque_002.xlsx, etc.")
        print("   Vérifiez que:")
        print("   - Les fichiers sont bien nommés avec le format Plaque_XXX.xlsx")
        print("   - Les numéros ont 3 chiffres (001, 002, etc.)")
        print("\n   Exemples de fichiers dans le dossier:")
        for f in tous_fichiers[:10]:
            print(f"   - {f}")
        return
    
    print(f"\n   Fichiers Plaque à traiter:")
    for f in fichiers_excel[:5]:
        print(f"   - {f}")
    if len(fichiers_excel) > 5:
        print(f"   ... et {len(fichiers_excel) - 5} autres")
    print(f"\n   Configuration: Tray ID + {nb_dates_photo} colonnes de dates photo")
    print()
    
    # Créer le dossier de destination s'il n'existe pas
    if dossier_destination is None:
        dossier_destination = os.path.join(dossier_source, "modifiés")
    
    if not os.path.exists(dossier_destination):
        os.makedirs(dossier_destination)
    
    # Compter les fichiers traités
    fichiers_traites = 0
    fichiers_erreur = 0
    
    # Parcourir tous les fichiers Excel
    for fichier in fichiers_excel:
        chemin_fichier = os.path.join(dossier_source, fichier)
        
        try:
            print(f"Traitement de: {fichier}...")
            
            # Charger le fichier Excel
            wb = load_workbook(chemin_fichier)
            
            # Afficher les feuilles disponibles
            print(f"   Feuilles disponibles: {', '.join(wb.sheetnames)}")
            
            # Vérifier si la feuille 'Suivi' existe
            if 'Suivi' not in wb.sheetnames:
                print(f"  ⚠️  Feuille 'Suivi' non trouvée dans {fichier}")
                print(f"      Utilisez une de ces feuilles: {', '.join(wb.sheetnames)}")
                fichiers_erreur += 1
                continue
            
            # Sélectionner la feuille 'Suivi'
            ws = wb['Suivi']
            
            # Trouver la dernière colonne utilisée
            derniere_colonne = ws.max_column
            
            # Copier le style de la dernière colonne existante (en-tête)
            cellule_reference = ws.cell(row=1, column=derniere_colonne)
            
            # Ajouter la colonne 'Tray ID'
            nouvelle_col = derniere_colonne + 1
            cellule_tray = ws.cell(row=1, column=nouvelle_col, value='Tray ID')
            
            # Copier le style
            if cellule_reference.font:
                cellule_tray.font = copy(cellule_reference.font)
            if cellule_reference.border:
                cellule_tray.border = copy(cellule_reference.border)
            if cellule_reference.fill:
                cellule_tray.fill = copy(cellule_reference.fill)
            if cellule_reference.alignment:
                cellule_tray.alignment = copy(cellule_reference.alignment)
            
            # Ajouter les colonnes de dates photo
            for i in range(1, nb_dates_photo + 1):
                col_index = nouvelle_col + i
                cellule_date = ws.cell(row=1, column=col_index, value=f'Date photo {i}')
                
                # Copier le style
                if cellule_reference.font:
                    cellule_date.font = copy(cellule_reference.font)
                if cellule_reference.border:
                    cellule_date.border = copy(cellule_reference.border)
                if cellule_reference.fill:
                    cellule_date.fill = copy(cellule_reference.fill)
                if cellule_reference.alignment:
                    cellule_date.alignment = copy(cellule_reference.alignment)
            
            # Ajuster la largeur des nouvelles colonnes (optionnel)
            for col in range(nouvelle_col, nouvelle_col + nb_dates_photo + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
            # Sauvegarder le fichier modifié
            chemin_destination = os.path.join(dossier_destination, fichier)
            wb.save(chemin_destination)
            
            print(f"  ✓ {fichier} traité avec succès")
            fichiers_traites += 1
            
        except Exception as e:
            print(f"  ✗ Erreur avec {fichier}: {str(e)}")
            fichiers_erreur += 1
    
    # Résumé
    print("\n" + "="*50)
    print(f"Traitement terminé !")
    print(f"✓ Fichiers traités: {fichiers_traites}")
    print(f"✗ Fichiers en erreur: {fichiers_erreur}")
    print(f"📁 Fichiers sauvegardés dans: {dossier_destination}")
    print("="*50)


if __name__ == "__main__":
    # MODIFIEZ CE CHEMIN AVEC LE DOSSIER CONTENANT VOS FICHIERS
    dossier_source = r"C:\IE\Etudes\ET_Corélac\CORELAC_300_plaques_24_femelles_groupées"
    
    # MODIFIEZ LE NOMBRE DE COLONNES POUR LES DATES PHOTO (défaut: 5)
    nb_dates_photo = 50  # Changez ce nombre selon vos besoins
    
    # Optionnel: spécifier un dossier de destination différent
    # dossier_destination = r"C:\chemin\vers\destination"
    
    print("Début du traitement des fichiers Excel...")
    print(f"Dossier source: {dossier_source}\n")
    
    ajouter_colonnes_excel(dossier_source, nb_dates_photo=nb_dates_photo)
    
    print("\nAppuyez sur Entrée pour fermer...")
    input()