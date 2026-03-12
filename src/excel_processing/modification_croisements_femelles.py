import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy

def modifier_croisements_plaques(dossier_source, dossier_destination=None):
    """
    Modifie les croisements dans les fichiers Plaque_001 à Plaque_200.
    
    Modifications appliquées:
    1. Remplacement des femelles B_F11 à B_F15 selon nouvelle matrice (en gras)
    2. Interversion B_M1 ↔ B_M7 pour croisements BOURGET×BOURGET (en italique)
    
    Args:
        dossier_source: Chemin du dossier contenant les fichiers Excel
        dossier_destination: Chemin du dossier pour les fichiers modifiés
    """
    
    # Dictionnaire de remplacement pour les femelles B_F11 à B_F15
    # Format: {ancien_croisement: nouveau_croisement}
    remplacements_femelles = {
        # Bourget × Bourget - Mâles M11, M12, M13
        'B_M11xB_F11': 'B_M11xB_F1', 'B_M11xB_F12': 'B_M11xB_F2', 'B_M11xB_F13': 'B_M11xB_F3',
        'B_M11xB_F14': 'B_M11xB_F4', 'B_M11xB_F15': 'B_M11xB_F5',
        
        'B_M12xB_F11': 'B_M12xB_F1', 'B_M12xB_F12': 'B_M12xB_F2', 'B_M12xB_F13': 'B_M12xB_F3',
        'B_M12xB_F14': 'B_M12xB_F4', 'B_M12xB_F15': 'B_M12xB_F5',
        
        'B_M13xB_F11': 'B_M13xB_F1', 'B_M13xB_F12': 'B_M13xB_F2', 'B_M13xB_F13': 'B_M13xB_F3',
        'B_M13xB_F14': 'B_M13xB_F4', 'B_M13xB_F15': 'B_M13xB_F5',
        
        # Bourget × Bourget - Mâles M14, M15
        'B_M14xB_F11': 'B_M14xB_F6', 'B_M14xB_F12': 'B_M14xB_F7', 'B_M14xB_F13': 'B_M14xB_F8',
        'B_M14xB_F14': 'B_M14xB_F9', 'B_M14xB_F15': 'B_M14xB_F10',
        
        'B_M15xB_F11': 'B_M15xB_F6', 'B_M15xB_F12': 'B_M15xB_F7', 'B_M15xB_F13': 'B_M15xB_F8',
        'B_M15xB_F14': 'B_M15xB_F9', 'B_M15xB_F15': 'B_M15xB_F10',
        
        # Léman × Bourget - Mâles M11, M12, M13
        'L_M11xB_F11': 'L_M11xB_F1', 'L_M11xB_F12': 'L_M11xB_F2', 'L_M11xB_F13': 'L_M11xB_F3',
        'L_M11xB_F14': 'L_M11xB_F4', 'L_M11xB_F15': 'L_M11xB_F5',
        
        'L_M12xB_F11': 'L_M12xB_F1', 'L_M12xB_F12': 'L_M12xB_F2', 'L_M12xB_F13': 'L_M12xB_F3',
        'L_M12xB_F14': 'L_M12xB_F4', 'L_M12xB_F15': 'L_M12xB_F5',
        
        'L_M13xB_F11': 'L_M13xB_F1', 'L_M13xB_F12': 'L_M13xB_F2', 'L_M13xB_F13': 'L_M13xB_F3',
        'L_M13xB_F14': 'L_M13xB_F4', 'L_M13xB_F15': 'L_M13xB_F5',
        
        # Léman × Bourget - Mâles M14, M15
        'L_M14xB_F11': 'L_M14xB_F6', 'L_M14xB_F12': 'L_M14xB_F7', 'L_M14xB_F13': 'L_M14xB_F8',
        'L_M14xB_F14': 'L_M14xB_F9', 'L_M14xB_F15': 'L_M14xB_F10',
        
        'L_M15xB_F11': 'L_M15xB_F6', 'L_M15xB_F12': 'L_M15xB_F7', 'L_M15xB_F13': 'L_M15xB_F8',
        'L_M15xB_F14': 'L_M15xB_F9', 'L_M15xB_F15': 'L_M15xB_F10',
    }
    
    # Vérifier que le dossier source existe
    if not os.path.exists(dossier_source):
        print(f"❌ ERREUR: Le dossier '{dossier_source}' n'existe pas!")
        return
    
    # Lister tous les fichiers Plaque_XXX
    tous_fichiers = os.listdir(dossier_source)
    fichiers_excel = []
    for f in tous_fichiers:
        if re.match(r'Plaque_\d{3}\.(xlsx|xls)$', f, re.IGNORECASE):
            fichiers_excel.append(f)
    
    fichiers_excel.sort()
    
    print(f"📁 Fichiers Plaque_XXX trouvés: {len(fichiers_excel)}")
    
    if len(fichiers_excel) == 0:
        print("\n⚠️  Aucun fichier Plaque_XXX.xlsx trouvé!")
        return
    
    # Créer le dossier de destination
    if dossier_destination is None:
        dossier_destination = os.path.join(dossier_source, "plaques_modifiées")
    
    if not os.path.exists(dossier_destination):
        os.makedirs(dossier_destination)
    
    print("\n🔧 Modifications à appliquer:")
    print("  1. Remplacement femelles B_F11-F15 → nouvelles femelles (GRAS)")
    print("  2. Interversion B_M1 ↔ B_M7 pour BOURGET×BOURGET (ITALIQUE)")
    print()
    
    fichiers_traites = 0
    total_modifs_femelles = 0
    total_interversions = 0
    
    for fichier in fichiers_excel:
        chemin_fichier = os.path.join(dossier_source, fichier)
        nom_plaque = os.path.splitext(fichier)[0]
        
        try:
            print(f"Traitement: {fichier}...")
            
            # Charger le fichier Excel
            wb = load_workbook(chemin_fichier)
            ws = wb.worksheets[0]  # Première feuille
            
            modifs_femelles_plaque = 0
            interversions_plaque = 0
            
            # Parcourir toutes les cellules
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        valeur_originale = cell.value
                        nouvelle_valeur = valeur_originale
                        appliquer_gras = False
                        appliquer_italique = False
                        
                        # Modification 1: Remplacement des femelles B_F11-F15
                        if valeur_originale in remplacements_femelles:
                            nouvelle_valeur = remplacements_femelles[valeur_originale]
                            appliquer_gras = True
                            modifs_femelles_plaque += 1
                        
                        # Modification 2: Interversion B_M1 ↔ B_M7 (uniquement BOURGET×BOURGET)
                        # Vérifier si c'est un croisement BOURGET×BOURGET avec M1 ou M7
                        if nouvelle_valeur.startswith('B_M1x') and 'B_F' in nouvelle_valeur:
                            # Remplacer B_M1 par B_M7
                            nouvelle_valeur = nouvelle_valeur.replace('B_M1x', 'B_M7x')
                            appliquer_italique = True
                            interversions_plaque += 1
                        elif nouvelle_valeur.startswith('B_M7x') and 'B_F' in nouvelle_valeur:
                            # Remplacer B_M7 par B_M1
                            nouvelle_valeur = nouvelle_valeur.replace('B_M7x', 'B_M1x')
                            appliquer_italique = True
                            interversions_plaque += 1
                        
                        # Appliquer les modifications
                        if nouvelle_valeur != valeur_originale or appliquer_gras or appliquer_italique:
                            cell.value = nouvelle_valeur
                            
                            # Copier le style existant ou créer un nouveau
                            if cell.font:
                                nouvelle_font = copy(cell.font)
                            else:
                                nouvelle_font = Font()
                            
                            # Appliquer le formatage
                            if appliquer_gras:
                                nouvelle_font = Font(
                                    name=nouvelle_font.name,
                                    size=nouvelle_font.size,
                                    bold=True,
                                    italic=nouvelle_font.italic or appliquer_italique,
                                    color=nouvelle_font.color
                                )
                            if appliquer_italique:
                                nouvelle_font = Font(
                                    name=nouvelle_font.name,
                                    size=nouvelle_font.size,
                                    bold=nouvelle_font.bold or appliquer_gras,
                                    italic=True,
                                    color=nouvelle_font.color
                                )
                            
                            cell.font = nouvelle_font
            
            # Sauvegarder le fichier modifié
            chemin_destination = os.path.join(dossier_destination, fichier)
            wb.save(chemin_destination)
            
            print(f"  ✓ Modifié: {modifs_femelles_plaque} remplacements femelles, {interversions_plaque} interversions")
            fichiers_traites += 1
            total_modifs_femelles += modifs_femelles_plaque
            total_interversions += interversions_plaque
            
        except Exception as e:
            print(f"  ✗ Erreur avec {fichier}: {str(e)}")
    
    # Résumé
    print("\n" + "="*60)
    print(f"Traitement terminé !")
    print(f"✓ Fichiers traités: {fichiers_traites}/{len(fichiers_excel)}")
    print(f"📊 Total modifications femelles (GRAS): {total_modifs_femelles}")
    print(f"🔄 Total interversions B_M1↔B_M7 (ITALIQUE): {total_interversions}")
    print(f"📁 Fichiers sauvegardés dans: {dossier_destination}")
    print("="*60)


if __name__ == "__main__":
    # MODIFIEZ CE CHEMIN AVEC LE DOSSIER CONTENANT VOS FICHIERS
    dossier_source = r"C:\IE\Etudes\ET_Corélac\CORELAC_300_plaques_24_femelles_groupées"
    
    # Optionnel: spécifier un dossier de destination différent
    # dossier_destination = r"C:\chemin\vers\destination"
    
    print("🔧 Début de la modification des croisements...")
    print(f"📁 Dossier source: {dossier_source}\n")
    
    modifier_croisements_plaques(dossier_source)
    
    print("\nAppuyez sur Entrée pour fermer...")
    input()